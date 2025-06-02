"""
Enhanced Workflow Matcher with Final Clean Matcher Integration
Clean, well-organized version with requirement quality analysis and full explainability
"""

import pandas as pd
import numpy as np
from pathlib import Path
import json
from datetime import datetime
import logging
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re

# Import the enhanced matcher and requirements analyzer
from matcher import FinalCleanMatcher, MatchExplanation
from reqGrading import RequirementAnalyzer, QualityMetrics

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class WorkflowMatchResult:
    """Enhanced match result for workflow processing with explainability and quality metrics."""
    requirement_id: str
    requirement_name: str
    requirement_text: str
    activity_name: str
    combined_score: float
    semantic_score: float
    bm25_score: float
    syntactic_score: float
    domain_score: float
    query_expansion_score: float
    confidence_level: str
    match_type: str
    review_priority: str
    explanation: MatchExplanation
    engineering_notes: str = ""
    review_status: str = "PENDING"
    # Quality metrics
    requirement_quality_score: float = 0.0
    clarity_score: float = 0.0
    completeness_score: float = 0.0
    verifiability_score: float = 0.0
    atomicity_score: float = 0.0
    consistency_score: float = 0.0
    quality_issues: List[str] = None
    quality_grade: str = "Unknown"

class EnhancedWorkflowMatcher(FinalCleanMatcher):
    """Enhanced workflow matcher with explainability and quality analysis for engineering teams."""
    
    def __init__(self, model_name: str = "en_core_web_trf"):
        super().__init__(model_name)
        self.confidence_thresholds = {
            'high': 0.7,
            'medium': 0.4,
            'low': 0.2
        }
        self.match_types = {
            'exact': 0.9,
            'strong': 0.7,
            'moderate': 0.4,
            'weak': 0.2
        }
        # Initialize requirement quality analyzer
        self.quality_analyzer = RequirementAnalyzer()
        self.requirement_quality_cache = {}
    
    # Quality Analysis Methods
    def analyze_requirement_quality(self, requirement_text: str, requirement_id: str) -> Tuple[QualityMetrics, List[str]]:
        """Analyze requirement quality and cache results."""
        if requirement_id in self.requirement_quality_cache:
            return self.requirement_quality_cache[requirement_id]
        
        issues, metrics = self.quality_analyzer.analyze_requirement(requirement_text, requirement_id)
        self.requirement_quality_cache[requirement_id] = (metrics, issues)
        return metrics, issues
    
    def determine_quality_grade(self, overall_score: float) -> str:
        """Determine quality grade based on overall score."""
        if overall_score >= 80:
            return "EXCELLENT"
        elif overall_score >= 65:
            return "GOOD"
        elif overall_score >= 50:
            return "FAIR"
        elif overall_score >= 35:
            return "POOR"
        else:
            return "CRITICAL"
    
    def explain_poor_match_with_quality(self, match_score: float, quality_score: float, 
                                       quality_issues: List[str]) -> str:
        """Generate explanation for poor matches considering requirement quality."""
        explanations = []
        
        if match_score < 0.4 and quality_score < 50:
            explanations.append("Low match score likely due to poor requirement quality")
            
        if quality_score < 35:
            explanations.append("CRITICAL quality issues may prevent accurate matching")
            
        # Identify specific quality issues that affect matching
        matching_affecting_issues = [
            "ambiguous term", "no measurable criteria", "missing subject", 
            "missing verb", "missing object", "no modal verb", "vague phrase"
        ]
        
        relevant_issues = []
        for issue in quality_issues:
            if any(affecting in issue.lower() for affecting in matching_affecting_issues):
                relevant_issues.append(issue)
        
        if relevant_issues:
            explanations.append(f"Specific issues affecting matching: {'; '.join(relevant_issues[:3])}")
            
        return " | ".join(explanations) if explanations else "Quality analysis available"

    # Match Classification Methods
    def categorize_match(self, scores: Dict[str, float]) -> Tuple[str, str, str]:
        """Categorize match quality and determine review priority."""
        combined = scores.get('combined', 0)
        semantic = scores.get('dense_semantic', 0)
        
        # Determine confidence level
        if combined >= self.confidence_thresholds['high']:
            confidence = 'HIGH'
        elif combined >= self.confidence_thresholds['medium']:
            confidence = 'MEDIUM'
        else:
            confidence = 'LOW'
        
        # Determine match type
        if combined >= self.match_types['exact']:
            match_type = 'EXACT'
        elif combined >= self.match_types['strong']:
            match_type = 'STRONG'
        elif combined >= self.match_types['moderate']:
            match_type = 'MODERATE'
        else:
            match_type = 'WEAK'
        
        # Determine review priority based on confidence and semantic strength
        if confidence == 'HIGH' and semantic > 0.6:
            priority = 'AUTO-APPROVE'
        elif confidence == 'HIGH' or (confidence == 'MEDIUM' and semantic > 0.5):
            priority = 'QUICK-REVIEW'
        elif confidence == 'MEDIUM':
            priority = 'DETAILED-REVIEW'
        else:
            priority = 'MANUAL-ANALYSIS'
        
        return confidence, match_type, priority
    
    # Workflow Processing Methods
    def generate_workflow_matches(self, matches_df: pd.DataFrame, 
                                explanations: List[MatchExplanation]) -> List[WorkflowMatchResult]:
        """Convert DataFrame and explanations to workflow match results with quality analysis."""
        workflow_matches = []
        seen_pairs = set()  # Track requirement-activity pairs to avoid duplicates
        
        # Create lookup for explanations by requirement+activity
        explanation_lookup = {}
        for exp in explanations:
            key = f"{exp.requirement_id}||{exp.activity_name}"
            explanation_lookup[key] = exp
        
        logger.info("Analyzing requirement quality for workflow matches...")
        
        for _, row in matches_df.iterrows():
            # Create unique pair identifier
            pair_key = (str(row['ID']), str(row['Activity Name']))
            
            # Skip if we've already seen this requirement-activity pair
            if pair_key in seen_pairs:
                continue
            seen_pairs.add(pair_key)
            
            # Analyze requirement quality
            req_id = str(row['ID'])
            req_text = str(row['Requirement Text'])
            quality_metrics, quality_issues = self.analyze_requirement_quality(req_text, req_id)
            
            # Calculate overall quality score
            overall_quality = (
                quality_metrics.clarity_score * 0.25 +
                quality_metrics.completeness_score * 0.25 +
                quality_metrics.verifiability_score * 0.2 +
                quality_metrics.atomicity_score * 0.2 +
                quality_metrics.consistency_score * 0.1
            )
            
            quality_grade = self.determine_quality_grade(overall_quality)
            
            # Get explanation for this match
            lookup_key = f"{row['ID']}||{row['Activity Name']}"
            explanation = explanation_lookup.get(lookup_key)
            
            if not explanation:
                # Create a basic explanation if not found
                explanation = MatchExplanation(
                    requirement_id=str(row['ID']),
                    requirement_text=str(row['Requirement Text']),
                    activity_name=str(row['Activity Name']),
                    combined_score=row['Combined Score'],
                    semantic_score=row.get('Dense Semantic', 0),
                    bm25_score=row.get('BM25 Score', 0),
                    syntactic_score=row.get('Syntactic Score', 0),
                    domain_score=row.get('Domain Weighted', 0),
                    query_expansion_score=row.get('Query Expansion', 0),
                    semantic_explanation="Generated from match data",
                    bm25_explanation="Generated from match data",
                    syntactic_explanation="Generated from match data",
                    domain_explanation="Generated from match data",
                    query_expansion_explanation="Generated from match data",
                    shared_terms=[],
                    semantic_similarity_level="Unknown",
                    match_quality="Unknown"
                )
            
            scores = {
                'combined': row['Combined Score'],
                'dense_semantic': row.get('Dense Semantic', 0),
                'bm25': row.get('BM25 Score', 0),
                'syntactic': row.get('Syntactic Score', 0),
                'domain_weighted': row.get('Domain Weighted', 0),
                'query_expansion': row.get('Query Expansion', 0)
            }
            
            confidence, match_type, priority = self.categorize_match(scores)
            
            # Enhance explanation with quality analysis for poor matches
            if row['Combined Score'] < 0.5:
                quality_explanation = self.explain_poor_match_with_quality(
                    row['Combined Score'], overall_quality, quality_issues
                )
                if quality_explanation and quality_explanation != "Quality analysis available":
                    explanation.semantic_explanation += f" | Quality Impact: {quality_explanation}"
            
            workflow_match = WorkflowMatchResult(
                requirement_id=str(row['ID']),
                requirement_name=str(row.get('Requirement Name', '')),
                requirement_text=str(row['Requirement Text']),
                activity_name=str(row['Activity Name']),
                combined_score=row['Combined Score'],
                semantic_score=scores['dense_semantic'],
                bm25_score=scores['bm25'],
                syntactic_score=scores['syntactic'],
                domain_score=scores['domain_weighted'],
                query_expansion_score=scores['query_expansion'],
                confidence_level=confidence,
                match_type=match_type,
                review_priority=priority,
                explanation=explanation,
                # Quality metrics
                requirement_quality_score=overall_quality,
                clarity_score=quality_metrics.clarity_score,
                completeness_score=quality_metrics.completeness_score,
                verifiability_score=quality_metrics.verifiability_score,
                atomicity_score=quality_metrics.atomicity_score,
                consistency_score=quality_metrics.consistency_score,
                quality_issues=quality_issues,
                quality_grade=quality_grade
            )
            workflow_matches.append(workflow_match)
        
        logger.info(f"Generated {len(workflow_matches)} unique workflow matches from {len(matches_df)} total matches")
        logger.info(f"Quality analysis: {len([m for m in workflow_matches if m.quality_grade in ['POOR', 'CRITICAL']])} requirements need quality improvement")
        return workflow_matches

    # Excel Generation Methods
    def create_engineering_review_package(self, matches_df: pd.DataFrame, 
                                        explanations: List[MatchExplanation],
                                        output_dir: str = "engineering_review") -> Dict[str, str]:
        """Create complete package for internal engineering review with explanations."""
        
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        workflow_matches = self.generate_workflow_matches(matches_df, explanations)
        files_created = {}
        
        # 1. Enhanced Excel workbook with explanations
        excel_file = output_path / "dependency_review_workbook_explained.xlsx"
        self._create_enhanced_engineering_excel(workflow_matches, excel_file)
        files_created['review_workbook'] = str(excel_file)
        
        # 2. Structured JSON with full explanations
        json_file = output_path / "matches_with_explanations.json"
        self._create_structured_json_with_explanations(workflow_matches, json_file)
        files_created['structured_data'] = str(json_file)
        
        # 3. Enhanced action items with explanation summaries
        action_file = output_path / "action_items_explained.csv"
        self._create_enhanced_action_items(workflow_matches, action_file)
        files_created['action_items'] = str(action_file)
        
        # 4. Explanation summary report
        explanation_report = output_path / "explanation_summary.html"
        self._create_explanation_summary_html(workflow_matches, explanation_report)
        files_created['explanation_report'] = str(explanation_report)
        
        # 5. Summary dashboard with explanation metrics
        summary_file = output_path / "match_summary_enhanced.json"
        self._create_enhanced_summary_dashboard(workflow_matches, summary_file)
        files_created['summary_dashboard'] = str(summary_file)
        
        # 6. Quick reference guide
        guide_file = output_path / "workflow_guide_quick_reference.md"
        self._create_quick_reference_guide(guide_file)
        files_created['workflow_guide'] = str(guide_file)
        
        logger.info(f"Created enhanced engineering review package in {output_dir}")
        return files_created
    
    def _create_enhanced_engineering_excel(self, matches: List[WorkflowMatchResult], filepath: Path):
        """Create comprehensive Excel workbook with explanations for engineering review."""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Sheet 1: Auto-Approve Candidates
        auto_approve = [m for m in matches if m.review_priority == 'AUTO-APPROVE']
        if auto_approve:
            ws1 = wb.create_sheet("Auto-Approve")
            self._populate_enhanced_excel_sheet(ws1, auto_approve, header_fill, header_font, high_fill)
        
        # Sheet 2: Quick Review Required
        quick_review = [m for m in matches if m.review_priority == 'QUICK-REVIEW']
        if quick_review:
            ws2 = wb.create_sheet("Quick Review")
            self._populate_enhanced_excel_sheet(ws2, quick_review, header_fill, header_font, medium_fill)
        
        # Sheet 3: Detailed Review Required
        detailed_review = [m for m in matches if m.review_priority == 'DETAILED-REVIEW']
        if detailed_review:
            ws3 = wb.create_sheet("Detailed Review")
            self._populate_enhanced_excel_sheet(ws3, detailed_review, header_fill, header_font, medium_fill)
        
        # Sheet 4: Manual Analysis Needed
        manual_analysis = [m for m in matches if m.review_priority == 'MANUAL-ANALYSIS']
        if manual_analysis:
            ws4 = wb.create_sheet("Manual Analysis")
            self._populate_enhanced_excel_sheet(ws4, manual_analysis, header_fill, header_font, low_fill)
        
        # Sheet 5: Enhanced Summary with Explanations
        ws_summary = wb.create_sheet("Summary")
        self._create_enhanced_summary_sheet(ws_summary, matches, header_fill, header_font)
        
        # Sheet 6: Explanation Guide
        ws_guide = wb.create_sheet("Explanation Guide")
        self._create_explanation_guide_sheet(ws_guide, header_fill, header_font)
        
        wb.save(filepath)
    
    def _populate_enhanced_excel_sheet(self, ws, matches: List[WorkflowMatchResult], 
                                     header_fill, header_font, row_fill):
        """Populate Excel sheet with enhanced match data, explanations, and quality analysis."""
        headers = [
            "Req ID", "Requirement Name", "Requirement Text", "Activity Name", "Combined Score",
            "Confidence", "Match Type", "Priority", "Semantic Score", "BM25 Score",
            "Syntactic Score", "Domain Score", "Query Expansion", "Match Quality",
            "Quality Grade", "Quality Score", "Clarity", "Completeness", "Verifiability", 
            "Atomicity", "Consistency", "Quality Issues", "Semantic Explanation", 
            "BM25 Explanation", "Domain Explanation", "Quality Impact", "Shared Terms", 
            "Engineer Review", "Status", "Notes"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Remove duplicates while preserving order
        seen_pairs = set()
        unique_matches = []
        for match in matches:
            pair_key = (match.requirement_id, match.activity_name)
            if pair_key not in seen_pairs:
                seen_pairs.add(pair_key)
                unique_matches.append(match)
        
        # Add data with full text, explanations, and quality analysis
        for row, match in enumerate(unique_matches, 2):
            ws.cell(row=row, column=1, value=match.requirement_id)
            ws.cell(row=row, column=2, value=match.requirement_name)
            ws.cell(row=row, column=3, value=match.requirement_text)  # Full text
            ws.cell(row=row, column=4, value=match.activity_name)  # Full text
            ws.cell(row=row, column=5, value=round(match.combined_score, 3))
            ws.cell(row=row, column=6, value=match.confidence_level)
            ws.cell(row=row, column=7, value=match.match_type)
            ws.cell(row=row, column=8, value=match.review_priority)
            ws.cell(row=row, column=9, value=round(match.semantic_score, 3))
            ws.cell(row=row, column=10, value=round(match.bm25_score, 3))
            ws.cell(row=row, column=11, value=round(match.syntactic_score, 3))
            ws.cell(row=row, column=12, value=round(match.domain_score, 3))
            ws.cell(row=row, column=13, value=round(match.query_expansion_score, 3))
            ws.cell(row=row, column=14, value=match.explanation.match_quality)
            # Quality metrics
            ws.cell(row=row, column=15, value=match.quality_grade)
            ws.cell(row=row, column=16, value=round(match.requirement_quality_score, 1))
            ws.cell(row=row, column=17, value=round(match.clarity_score, 1))
            ws.cell(row=row, column=18, value=round(match.completeness_score, 1))
            ws.cell(row=row, column=19, value=round(match.verifiability_score, 1))
            ws.cell(row=row, column=20, value=round(match.atomicity_score, 1))
            ws.cell(row=row, column=21, value=round(match.consistency_score, 1))
            ws.cell(row=row, column=22, value="; ".join(match.quality_issues[:3]) if match.quality_issues else "None")
            # Explanations
            ws.cell(row=row, column=23, value=match.explanation.semantic_explanation)  # Full explanation
            ws.cell(row=row, column=24, value=match.explanation.bm25_explanation)  # Full explanation
            ws.cell(row=row, column=25, value=match.explanation.domain_explanation)  # Full explanation
            ws.cell(row=row, column=26, value=self.explain_poor_match_with_quality(
                match.combined_score, match.requirement_quality_score, match.quality_issues))
            ws.cell(row=row, column=27, value=", ".join(match.explanation.shared_terms))
            ws.cell(row=row, column=28, value="")  # Engineer Review (to be filled)
            ws.cell(row=row, column=29, value="PENDING")  # Status
            ws.cell(row=row, column=30, value="")  # Notes
            
            # Apply row fill with quality-based coloring
            quality_fill = row_fill
            if match.quality_grade in ['POOR', 'CRITICAL']:
                quality_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")  # Light red
            elif match.quality_grade == 'FAIR':
                quality_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")  # Light yellow
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = quality_fill
        
        # Set optimal column widths for readability including quality columns
        column_widths = {
            'A': 12,  # Req ID
            'B': 25,  # Requirement Name
            'C': 80,  # Requirement Text (wide for full text)
            'D': 50,  # Activity Name
            'E': 15,  # Combined Score
            'F': 12,  # Confidence
            'G': 12,  # Match Type
            'H': 18,  # Priority
            'I': 12,  # Semantic Score
            'J': 12,  # BM25 Score
            'K': 12,  # Syntactic Score
            'L': 12,  # Domain Score
            'M': 15,  # Query Expansion
            'N': 15,  # Match Quality
            'O': 15,  # Quality Grade
            'P': 12,  # Quality Score
            'Q': 12,  # Clarity
            'R': 15,  # Completeness
            'S': 15,  # Verifiability
            'T': 12,  # Atomicity
            'U': 12,  # Consistency
            'V': 60,  # Quality Issues
            'W': 60,  # Semantic Explanation
            'X': 60,  # BM25 Explanation
            'Y': 60,  # Domain Explanation
            'Z': 60,  # Quality Impact
            'AA': 40, # Shared Terms
            'AB': 20, # Engineer Review
            'AC': 15, # Status
            'AD': 40  # Notes
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Enable text wrapping for text-heavy columns
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        for row in range(2, len(unique_matches) + 2):
            # Apply text wrapping to text-heavy columns
            for col in ['C', 'D', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AD']:  # Text columns
                ws[f'{col}{row}'].alignment = wrap_alignment
    
    def _create_enhanced_summary_sheet(self, ws, matches: List[WorkflowMatchResult], 
                                     header_fill, header_font):
        """Create enhanced summary statistics sheet with explanation metrics."""
        ws.title = "Summary"
        
        # Summary statistics
        total_matches = len(matches)
        by_confidence = {}
        by_priority = {}
        by_match_quality = {}
        by_req_quality = {}
        
        for match in matches:
            by_confidence[match.confidence_level] = by_confidence.get(match.confidence_level, 0) + 1
            by_priority[match.review_priority] = by_priority.get(match.review_priority, 0) + 1
            by_match_quality[match.explanation.match_quality] = by_match_quality.get(match.explanation.match_quality, 0) + 1
            by_req_quality[match.quality_grade] = by_req_quality.get(match.quality_grade, 0) + 1
        
        # Write summary data
        row = 1
        ws.cell(row=row, column=1, value="ENHANCED MATCH ANALYSIS SUMMARY").font = Font(size=16, bold=True)
        row += 2
        
        ws.cell(row=row, column=1, value="Total Matches:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_matches)
        row += 2
        
        ws.cell(row=row, column=1, value="By Confidence Level:").font = Font(bold=True)
        row += 1
        for conf, count in by_confidence.items():
            ws.cell(row=row, column=2, value=f"{conf}: {count} ({count/total_matches*100:.1f}%)")
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="By Review Priority:").font = Font(bold=True)
        row += 1
        for priority, count in by_priority.items():
            estimated_hours = count * self._estimate_review_hours(priority)
            ws.cell(row=row, column=2, value=f"{priority}: {count} ({estimated_hours:.1f} hours)")
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="By Requirement Quality:").font = Font(bold=True)
        row += 1
        for quality, count in by_req_quality.items():
            ws.cell(row=row, column=2, value=f"{quality}: {count} ({count/total_matches*100:.1f}%)")
            row += 1
    
    def _create_explanation_guide_sheet(self, ws, header_fill, header_font):
        """Create explanation guide sheet."""
        ws.title = "Explanation Guide"
        
        guide_content = [
            ["SCORING COMPONENT EXPLANATIONS", ""],
            ["", ""],
            ["Semantic Score", "Neural network-based meaning similarity (0-1)"],
            ["", "Very High (>0.7): Strong conceptual match"],
            ["", "High (0.5-0.7): Good conceptual alignment"],
            ["", "Medium (0.3-0.5): Moderate conceptual similarity"],
            ["", "Low (<0.3): Weak conceptual connection"],
            ["", ""],
            ["BM25 Score", "Statistical relevance ranking based on term frequency"],
            ["", "Shows exact term matches and their importance"],
            ["", "Higher scores indicate more shared technical terms"],
            ["", ""],
            ["Quality Score", "Requirement quality analysis (0-100)"],
            ["", "EXCELLENT (80+): Well-written, clear requirements"],
            ["", "GOOD (65-79): Good quality with minor issues"],
            ["", "FAIR (50-64): Acceptable but needs improvement"],
            ["", "POOR (35-49): Significant quality issues"],
            ["", "CRITICAL (<35): Major rewrite needed"],
            ["", ""],
            ["CONFIDENCE LEVELS", ""],
            ["", ""],
            ["HIGH (>0.7)", "Strong match, likely correct dependency"],
            ["MEDIUM (0.4-0.7)", "Moderate match, requires review"],
            ["LOW (<0.4)", "Weak match, may be false positive"],
            ["", ""],
            ["QUALITY IMPACT", ""],
            ["", "Poor quality requirements often lead to poor matches"],
            ["", "Focus on improving requirement quality for better results"],
            ["", "Critical quality issues may prevent accurate matching"]
        ]
        
        for row_idx, (col1, col2) in enumerate(guide_content, 1):
            if col1.isupper() and col1.endswith("S"):  # Headers
                cell = ws.cell(row=row_idx, column=1, value=col1)
                cell.font = Font(bold=True, size=12)
            else:
                ws.cell(row=row_idx, column=1, value=col1)
                ws.cell(row=row_idx, column=2, value=col2)
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60

    # Output File Generation Methods
    def _create_structured_json_with_explanations(self, matches: List[WorkflowMatchResult], filepath: Path):
        """Create structured JSON with full explanations."""
        data = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "total_matches": len(matches),
                "tool_version": "2.0_with_explanations_and_quality",
                "confidence_thresholds": self.confidence_thresholds
            },
            "matches": []
        }
        
        for match in matches:
            match_data = {
                "requirement": {
                    "id": match.requirement_id,
                    "name": match.requirement_name,
                    "text": match.requirement_text
                },
                "activity": {
                    "name": match.activity_name
                },
                "scores": {
                    "combined": round(match.combined_score, 3),
                    "semantic": round(match.semantic_score, 3),
                    "bm25": round(match.bm25_score, 3),
                    "syntactic": round(match.syntactic_score, 3),
                    "domain": round(match.domain_score, 3),
                    "query_expansion": round(match.query_expansion_score, 3)
                },
                "classification": {
                    "confidence_level": match.confidence_level,
                    "match_type": match.match_type,
                    "review_priority": match.review_priority,
                    "match_quality": match.explanation.match_quality
                },
                "quality_analysis": {
                    "overall_score": round(match.requirement_quality_score, 1),
                    "grade": match.quality_grade,
                    "clarity": round(match.clarity_score, 1),
                    "completeness": round(match.completeness_score, 1),
                    "verifiability": round(match.verifiability_score, 1),
                    "atomicity": round(match.atomicity_score, 1),
                    "consistency": round(match.consistency_score, 1),
                    "issues": match.quality_issues or []
                },
                "explanations": {
                    "semantic": match.explanation.semantic_explanation,
                    "bm25": match.explanation.bm25_explanation,
                    "syntactic": match.explanation.syntactic_explanation,
                    "domain": match.explanation.domain_explanation,
                    "query_expansion": match.explanation.query_expansion_explanation,
                    "quality_impact": self.explain_poor_match_with_quality(
                        match.combined_score, match.requirement_quality_score, match.quality_issues
                    )
                },
                "evidence": {
                    "shared_terms": match.explanation.shared_terms,
                    "semantic_similarity_level": match.explanation.semantic_similarity_level
                },
                "workflow": {
                    "engineering_notes": match.engineering_notes,
                    "review_status": match.review_status,
                    "estimated_hours": self._estimate_review_hours(match.review_priority)
                }
            }
            data["matches"].append(match_data)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    
    def _create_enhanced_action_items(self, matches: List[WorkflowMatchResult], filepath: Path):
        """Create enhanced action items CSV with explanation summaries."""
        action_data = []
        
        for match in matches:
            # Create concise explanation summary for action tracking
            explanation_summary = f"Semantic: {match.explanation.semantic_similarity_level}"
            if match.explanation.shared_terms:
                explanation_summary += f" | Shared: {', '.join(match.explanation.shared_terms[:2])}"
            if match.explanation.match_quality != "Unknown":
                explanation_summary += f" | Quality: {match.explanation.match_quality}"
            
            action_data.append({
                "Requirement_ID": match.requirement_id,
                "Activity_Name": match.activity_name,
                "Priority": match.review_priority,
                "Confidence": match.confidence_level,
                "Combined_Score": round(match.combined_score, 3),
                "Match_Quality": match.explanation.match_quality,
                "Req_Quality_Grade": match.quality_grade,
                "Req_Quality_Score": round(match.requirement_quality_score, 1),
                "Quality_Issues": "; ".join(match.quality_issues[:2]) if match.quality_issues else "None",
                "Explanation_Summary": explanation_summary,
                "Shared_Terms": ", ".join(match.explanation.shared_terms[:3]),
                "Assigned_Engineer": "",  # To be filled
                "Due_Date": "",  # To be filled
                "Status": match.review_status,
                "Estimated_Hours": self._estimate_review_hours(match.review_priority),
                "Engineering_Notes": match.engineering_notes
            })
        
        df = pd.DataFrame(action_data)
        df.to_csv(filepath, index=False)
    
    def _create_explanation_summary_html(self, matches: List[WorkflowMatchResult], filepath: Path):
        """Create HTML summary report with explanations and quality analysis for ALL matches."""
        
        # Remove duplicates while preserving order
        seen_pairs = set()
        unique_matches = []
        for match in matches:
            pair_key = (match.requirement_id, match.activity_name)
            if pair_key not in seen_pairs:
                seen_pairs.add(pair_key)
                unique_matches.append(match)
        
        # Sort by combined score descending
        sorted_matches = sorted(unique_matches, key=lambda x: x.combined_score, reverse=True)
        
        # Quality statistics
        quality_stats = {
            'EXCELLENT': len([m for m in sorted_matches if m.quality_grade == 'EXCELLENT']),
            'GOOD': len([m for m in sorted_matches if m.quality_grade == 'GOOD']),
            'FAIR': len([m for m in sorted_matches if m.quality_grade == 'FAIR']),
            'POOR': len([m for m in sorted_matches if m.quality_grade == 'POOR']),
            'CRITICAL': len([m for m in sorted_matches if m.quality_grade == 'CRITICAL'])
        }
        
        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>Complete Dependency Analysis with Quality Assessment</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.6; }}
        .header {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
        .summary-stats {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0; }}
        .match {{ border: 1px solid #ddd; margin: 10px 0; padding: 15px; border-radius: 5px; page-break-inside: avoid; }}
        .match-header {{ font-weight: bold; color: #2c3e50; margin-bottom: 10px; }}
        .requirement-text {{ background-color: #f0f8ff; padding: 10px; border-radius: 3px; margin: 10px 0; font-style: italic; }}
        .activity-text {{ background-color: #f0fff0; padding: 10px; border-radius: 3px; margin: 10px 0; font-weight: bold; }}
        .score-breakdown {{ margin: 10px 0; background-color: #fafafa; padding: 10px; border-radius: 3px; }}
        .quality-section {{ background-color: #fff3cd; padding: 10px; border-radius: 3px; margin: 10px 0; border-left: 4px solid #ffc107; }}
        .explanation {{ background-color: #f1f3f4; padding: 10px; border-left: 4px solid #4CAF50; margin: 5px 0; }}
        .high-confidence {{ border-left-color: #4CAF50; }}
        .medium-confidence {{ border-left-color: #FF9800; }}
        .low-confidence {{ border-left-color: #f44336; }}
        .quality-excellent {{ background-color: #d4edda; }}
        .quality-good {{ background-color: #d1ecf1; }}
        .quality-fair {{ background-color: #fff3cd; }}
        .quality-poor {{ background-color: #f8d7da; }}
        .quality-critical {{ background-color: #f5c6cb; border: 2px solid #dc3545; }}
        .evidence {{ font-style: italic; color: #666; margin: 10px 0; }}
        .filter-controls {{ margin: 20px 0; padding: 15px; background-color: #e9ecef; border-radius: 5px; }}
        .filter-button {{ background-color: #007bff; color: white; border: none; padding: 8px 15px; margin: 5px; border-radius: 3px; cursor: pointer; }}
        .filter-button:hover {{ background-color: #0056b3; }}
        .filter-button.active {{ background-color: #28a745; }}
        .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin: 20px 0; }}
        .stat-card {{ background-color: white; padding: 15px; border-radius: 5px; border: 1px solid #ddd; text-align: center; }}
        .quality-badge {{ display: inline-block; padding: 4px 8px; border-radius: 3px; font-size: 0.8em; font-weight: bold; }}
        .badge-excellent {{ background-color: #28a745; color: white; }}
        .badge-good {{ background-color: #17a2b8; color: white; }}
        .badge-fair {{ background-color: #ffc107; color: black; }}
        .badge-poor {{ background-color: #fd7e14; color: white; }}
        .badge-critical {{ background-color: #dc3545; color: white; }}
        .quality-details {{ font-size: 0.9em; margin: 5px 0; }}
    </style>
    <script>
        function filterMatches(criteria, value) {{
            const matches = document.querySelectorAll('.match');
            const buttons = document.querySelectorAll('.filter-button');
            
            // Reset button states
            buttons.forEach(btn => btn.classList.remove('active'));
            
            if (value === 'all') {{
                matches.forEach(match => match.style.display = 'block');
                document.querySelector('[data-filter="all"]').classList.add('active');
            }} else {{
                matches.forEach(match => {{
                    const matchValue = match.dataset[criteria];
                    match.style.display = matchValue === value ? 'block' : 'none';
                }});
                document.querySelector(`[data-filter="${{value}}"]`).classList.add('active');
            }}
        }}
    </script>
</head>
<body>
    <div class="header">
        <h1>Complete Dependency Analysis with Quality Assessment</h1>
        <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p>Total Matches: {len(sorted_matches)} (duplicates removed)</p>
    </div>
    
    <div class="summary-stats">
        <h2>Summary Statistics</h2>
        <div class="stats-grid">
            <div class="stat-card">
                <h3>{len(sorted_matches)}</h3>
                <p>Total Unique Matches</p>
            </div>
            <div class="stat-card">
                <h3>{len(set(m.requirement_id for m in sorted_matches))}</h3>
                <p>Requirements Covered</p>
            </div>
            <div class="stat-card">
                <h3>{np.mean([m.combined_score for m in sorted_matches]):.3f}</h3>
                <p>Average Match Score</p>
            </div>
            <div class="stat-card">
                <h3>{np.mean([m.requirement_quality_score for m in sorted_matches]):.1f}</h3>
                <p>Average Quality Score</p>
            </div>
            <div class="stat-card">
                <h3>{len([m for m in sorted_matches if m.confidence_level == 'HIGH'])}</h3>
                <p>High Confidence Matches</p>
            </div>
            <div class="stat-card">
                <h3>{quality_stats['POOR'] + quality_stats['CRITICAL']}</h3>
                <p>Requirements Needing Quality Improvement</p>
            </div>
        </div>
        
        <h3>Requirement Quality Distribution:</h3>
        <div class="stats-grid">
            <div class="stat-card quality-excellent">
                <h3>{quality_stats['EXCELLENT']}</h3>
                <p>Excellent Quality</p>
            </div>
            <div class="stat-card quality-good">
                <h3>{quality_stats['GOOD']}</h3>
                <p>Good Quality</p>
            </div>
            <div class="stat-card quality-fair">
                <h3>{quality_stats['FAIR']}</h3>
                <p>Fair Quality</p>
            </div>
            <div class="stat-card quality-poor">
                <h3>{quality_stats['POOR']}</h3>
                <p>Poor Quality</p>
            </div>
            <div class="stat-card quality-critical">
                <h3>{quality_stats['CRITICAL']}</h3>
                <p>Critical Quality Issues</p>
            </div>
        </div>
    </div>
    
    <div class="filter-controls">
        <h3>Filter by Review Priority:</h3>
        <button class="filter-button active" data-filter="all" onclick="filterMatches('priority', 'all')">Show All</button>
        <button class="filter-button" data-filter="AUTO-APPROVE" onclick="filterMatches('priority', 'AUTO-APPROVE')">Auto-Approve ({len([m for m in sorted_matches if m.review_priority == 'AUTO-APPROVE'])})</button>
        <button class="filter-button" data-filter="QUICK-REVIEW" onclick="filterMatches('priority', 'QUICK-REVIEW')">Quick Review ({len([m for m in sorted_matches if m.review_priority == 'QUICK-REVIEW'])})</button>
        <button class="filter-button" data-filter="DETAILED-REVIEW" onclick="filterMatches('priority', 'DETAILED-REVIEW')">Detailed Review ({len([m for m in sorted_matches if m.review_priority == 'DETAILED-REVIEW'])})</button>
        <button class="filter-button" data-filter="MANUAL-ANALYSIS" onclick="filterMatches('priority', 'MANUAL-ANALYSIS')">Manual Analysis ({len([m for m in sorted_matches if m.review_priority == 'MANUAL-ANALYSIS'])})</button>
        
        <h3>Filter by Quality Grade:</h3>
        <button class="filter-button" data-filter="EXCELLENT" onclick="filterMatches('quality', 'EXCELLENT')">Excellent ({quality_stats['EXCELLENT']})</button>
        <button class="filter-button" data-filter="GOOD" onclick="filterMatches('quality', 'GOOD')">Good ({quality_stats['GOOD']})</button>
        <button class="filter-button" data-filter="FAIR" onclick="filterMatches('quality', 'FAIR')">Fair ({quality_stats['FAIR']})</button>
        <button class="filter-button" data-filter="POOR" onclick="filterMatches('quality', 'POOR')">Poor ({quality_stats['POOR']})</button>
        <button class="filter-button" data-filter="CRITICAL" onclick="filterMatches('quality', 'CRITICAL')">Critical ({quality_stats['CRITICAL']})</button>
    </div>
    
    <h2>All Matches with Quality Analysis and Detailed Explanations</h2>
"""
        
        # Show ALL matches with quality analysis
        for i, match in enumerate(sorted_matches, 1):
            confidence_class = match.confidence_level.lower() + "-confidence"
            quality_class = f"quality-{match.quality_grade.lower()}"
            quality_badge_class = f"badge-{match.quality_grade.lower()}"
            
            # Determine if quality is affecting the match
            quality_impact = ""
            if match.combined_score < 0.5 and match.requirement_quality_score < 60:
                quality_impact = "Poor match score may be due to requirement quality issues"
            elif match.requirement_quality_score < 40:
                quality_impact = "Critical quality issues may prevent accurate matching"
            
            html_content += f"""
    <div class="match {confidence_class} {quality_class}" id="match-{i}" data-priority="{match.review_priority}" data-quality="{match.quality_grade}">
        <div class="match-header">
            Match {i}: {match.requirement_id} → {match.activity_name}
            <span class="quality-badge {quality_badge_class}">{match.quality_grade}</span>
        </div>
        
        <div class="requirement-text">
            <strong>Requirement:</strong> {match.requirement_text}
        </div>
        
        <div class="activity-text">
            <strong>Activity:</strong> {match.activity_name}
        </div>
        
        <div class="score-breakdown">
            <strong>Match Score:</strong> {match.combined_score:.3f} 
            | <strong>Confidence:</strong> {match.confidence_level}
            | <strong>Priority:</strong> {match.review_priority}
            | <strong>Match Quality:</strong> {match.explanation.match_quality}
            <br>
            <strong>Component Scores:</strong> 
            Semantic: {match.semantic_score:.3f} | 
            BM25: {match.bm25_score:.3f} | 
            Syntactic: {match.syntactic_score:.3f} | 
            Domain: {match.domain_score:.3f} | 
            Query Expansion: {match.query_expansion_score:.3f}
        </div>
        
        <div class="quality-section">
            <strong>Requirement Quality Analysis:</strong>
            <div class="quality-details">
                <strong>Overall Quality Score:</strong> {match.requirement_quality_score:.1f}/100 
                | <strong>Grade:</strong> {match.quality_grade}
            </div>
            <div class="quality-details">
                <strong>Quality Breakdown:</strong> 
                Clarity: {match.clarity_score:.1f} | 
                Completeness: {match.completeness_score:.1f} | 
                Verifiability: {match.verifiability_score:.1f} | 
                Atomicity: {match.atomicity_score:.1f} | 
                Consistency: {match.consistency_score:.1f}
            </div>
            {f'<div class="quality-details"><strong>Quality Issues:</strong> {"; ".join(match.quality_issues[:3])}</div>' if match.quality_issues else ''}
            {f'<div class="quality-details" style="color: #d63384;"><strong>{quality_impact}</strong></div>' if quality_impact else ''}
        </div>
        
        <div class="explanation">
            <strong>Semantic Analysis:</strong> {match.explanation.semantic_explanation}
        </div>
        
        <div class="explanation">
            <strong>Term Matching (BM25):</strong> {match.explanation.bm25_explanation}
        </div>
        
        <div class="explanation">
            <strong>Structural Analysis:</strong> {match.explanation.syntactic_explanation}
        </div>
        
        <div class="explanation">
            <strong>Domain Analysis:</strong> {match.explanation.domain_explanation}
        </div>
        
        <div class="explanation">
            <strong>Query Expansion:</strong> {match.explanation.query_expansion_explanation}
        </div>
        
        {f'<div class="evidence">Shared Terms: {", ".join(match.explanation.shared_terms)}</div>' if match.explanation.shared_terms else '<div class="evidence">No shared terms identified</div>'}
        
        {f'<div class="evidence"><strong>Quality Impact Analysis:</strong> {self.explain_poor_match_with_quality(match.combined_score, match.requirement_quality_score, match.quality_issues)}</div>' if match.combined_score < 0.6 else ''}
    </div>
"""
        
        html_content += """
    <script>
        // Initialize with all matches shown
        filterMatches('priority', 'all');
    </script>
</body>
</html>
"""
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def _create_enhanced_summary_dashboard(self, matches: List[WorkflowMatchResult], filepath: Path):
        """Create enhanced summary data for dashboard visualization with quality metrics."""
        # Calculate quality impact insights
        poor_quality_reqs = [m for m in matches if m.requirement_quality_score < 50]
        poor_matches = [m for m in matches if m.combined_score < 0.4]
        poor_quality_and_match = [m for m in matches if m.requirement_quality_score < 50 and m.combined_score < 0.4]
        
        # Quality vs Match Score correlation analysis
        quality_match_correlation = {
            "high_quality_good_match": len([m for m in matches if m.requirement_quality_score >= 65 and m.combined_score >= 0.6]),
            "high_quality_poor_match": len([m for m in matches if m.requirement_quality_score >= 65 and m.combined_score < 0.6]),
            "poor_quality_good_match": len([m for m in matches if m.requirement_quality_score < 65 and m.combined_score >= 0.6]),
            "poor_quality_poor_match": len([m for m in matches if m.requirement_quality_score < 65 and m.combined_score < 0.6])
        }
        
        summary = {
            "overview": {
                "total_matches": len(matches),
                "high_confidence": len([m for m in matches if m.confidence_level == 'HIGH']),
                "medium_confidence": len([m for m in matches if m.confidence_level == 'MEDIUM']),
                "low_confidence": len([m for m in matches if m.confidence_level == 'LOW']),
                "avg_combined_score": np.mean([m.combined_score for m in matches]),
                "avg_semantic_score": np.mean([m.semantic_score for m in matches]),
                "avg_quality_score": np.mean([m.requirement_quality_score for m in matches]),
                "requirements_needing_quality_improvement": len([m for m in matches if m.quality_grade in ['POOR', 'CRITICAL']])
            },
            "workflow_distribution": {
                "auto_approve": len([m for m in matches if m.review_priority == 'AUTO-APPROVE']),
                "quick_review": len([m for m in matches if m.review_priority == 'QUICK-REVIEW']),
                "detailed_review": len([m for m in matches if m.review_priority == 'DETAILED-REVIEW']),
                "manual_analysis": len([m for m in matches if m.review_priority == 'MANUAL-ANALYSIS'])
            },
            "estimated_effort": {
                "total_hours": sum(self._estimate_review_hours(m.review_priority) for m in matches),
                "by_priority": {
                    priority: sum(self._estimate_review_hours(m.review_priority) 
                                for m in matches if m.review_priority == priority)
                    for priority in set(m.review_priority for m in matches)
                }
            },
            "quality_analysis": {
                "quality_distribution": {
                    grade: len([m for m in matches if m.quality_grade == grade])
                    for grade in ['EXCELLENT', 'GOOD', 'FAIR', 'POOR', 'CRITICAL']
                },
                "quality_match_correlation": quality_match_correlation,
                "quality_impact_insights": {
                    "poor_quality_requirements": len(poor_quality_reqs),
                    "poor_matches": len(poor_matches),
                    "poor_quality_causing_poor_matches": len(poor_quality_and_match),
                    "quality_impact_percentage": (len(poor_quality_and_match) / max(len(poor_matches), 1)) * 100
                },
                "avg_quality_by_grade": {
                    grade: np.mean([m.requirement_quality_score for m in matches if m.quality_grade == grade])
                    for grade in set(m.quality_grade for m in matches) if grade != 'Unknown'
                }
            },
            "score_distributions": {
                "semantic_scores": [m.semantic_score for m in matches],
                "bm25_scores": [m.bm25_score for m in matches],
                "domain_scores": [m.domain_score for m in matches],
                "combined_scores": [m.combined_score for m in matches],
                "quality_scores": [m.requirement_quality_score for m in matches]
            }
        }
        
        with open(filepath, 'w') as f:
            json.dump(summary, f, indent=2)
    
    def _create_quick_reference_guide(self, filepath: Path):
        """Create quick reference guide with quality information."""
        guide_content = """# Enhanced Dependency Analysis - Quick Reference Guide

## Quality Grades:
- **EXCELLENT (80+)**: Ready for matching and approval
- **GOOD (65-79)**: Minor issues, safe to proceed  
- **FAIR (50-64)**: Some improvements recommended
- **POOR (35-49)**: Significant issues, review carefully
- **CRITICAL (<35)**: Rewrite required before approval

## Quality-Enhanced Decision Guidelines:

### Approve if:
- Good match score + GOOD/EXCELLENT quality
- Multiple shared terms + well-written requirement
- Strong semantic similarity + no critical quality issues

### Review Further if:
- Good match + FAIR quality (may improve with enhancement)
- Mixed scoring signals + moderate quality issues
- Moderate similarity + vocabulary mismatch vs quality issue

### Reject/Rewrite if:
- CRITICAL quality grade (regardless of match score)
- POOR quality + low match (quality-caused mismatch)
- No shared terms + significant quality issues

## Quality Impact on Matching:
- **High Quality + Poor Match**: Algorithm/vocabulary limitation
- **Poor Quality + Poor Match**: Requirement needs rewriting
- **Poor Quality + Good Match**: Investigate for false positive
- **High Quality + Good Match**: High confidence approval

## File Descriptions:
- **Excel Workbook**: Main review interface with all data and explanations
- **HTML Report**: Visual summary with filtering capabilities
- **JSON Data**: Machine-readable format for integration
- **Action Items CSV**: Project management tracking
- **Summary Dashboard**: Analytics and metrics

For detailed workflow instructions, see the main workflow guide documentation.
"""
        
        with open(filepath, 'w') as f:
            f.write(guide_content)
    
    def _estimate_review_hours(self, priority: str) -> float:
        """Estimate review hours based on priority."""
        hours_map = {
            'AUTO-APPROVE': 0.25,
            'QUICK-REVIEW': 0.5,
            'DETAILED-REVIEW': 2.0,
            'MANUAL-ANALYSIS': 4.0
        }
        return hours_map.get(priority, 2.0)
    
    # Main Workflow Method
    def run_enhanced_workflow_matching(self, requirements_file: str = "requirements.csv",
                                     activities_file: str = "activities.csv",
                                     gold_pairs_file: Optional[str] = None,
                                     min_sim: float = 0.35,
                                     top_n: int = 5,
                                     normalize_scores: bool = True,
                                     out_file: str = "enhanced_workflow_matches") -> pd.DataFrame:
        """Run enhanced matching with workflow support and explanations."""
        
        print("Starting Enhanced Workflow Matching with Explanations...")
        print("="*60)
        
        # Step 1: Run the final clean matching to get matches and explanations
        print("1. Running enhanced matching with explainability...")
        matches_df = self.run_final_matching(
            requirements_file=requirements_file,
            activities_file=activities_file,
            gold_pairs_file=gold_pairs_file,
            min_sim=min_sim,
            top_n=top_n,
            out_file=out_file
        )
        
        if matches_df.empty:
            print("No matches found. Exiting.")
            return matches_df
        
        # Step 2: Load explanations from the JSON file created by run_final_matching
        explanations_file = f"results/{out_file}_explanations.json"
        explanations = []
        
        try:
            with open(explanations_file, 'r', encoding='utf-8') as f:
                explanations_data = json.load(f)
                
            for exp_data in explanations_data:
                explanation = MatchExplanation(
                    requirement_id=exp_data['requirement_id'],
                    requirement_text=exp_data['requirement_text'],
                    activity_name=exp_data['activity_name'],
                    combined_score=exp_data['combined_score'],
                    semantic_score=exp_data['scores']['semantic'],
                    bm25_score=exp_data['scores']['bm25'],
                    syntactic_score=exp_data['scores'].get('syntactic', 0),
                    domain_score=exp_data['scores'].get('domain', 0),
                    query_expansion_score=exp_data['scores'].get('query_expansion', 0),
                    semantic_explanation=exp_data['explanations']['semantic'],
                    bm25_explanation=exp_data['explanations']['bm25'],
                    syntactic_explanation=exp_data['explanations'].get('syntactic', ''),
                    domain_explanation=exp_data['explanations'].get('domain', ''),
                    query_expansion_explanation=exp_data['explanations'].get('query_expansion', ''),
                    shared_terms=exp_data['shared_terms'],
                    semantic_similarity_level=exp_data['semantic_similarity_level'],
                    match_quality=exp_data['match_quality']
                )
                explanations.append(explanation)
                
        except FileNotFoundError:
            print(f"Warning: Could not load explanations from {explanations_file}")
            explanations = []
        
        print(f"   Found {len(matches_df)} matches with {len(explanations)} explanations")
        
        # Step 3: Create enhanced engineering review package
        print("2. Creating enhanced engineering review package...")
        files_created = self.create_engineering_review_package(
            matches_df, 
            explanations,
            output_dir="enhanced_engineering_review"
        )
        
        # Step 4: Enhanced summary for team
        print("\n" + "="*70)
        print("ENHANCED ENGINEERING REVIEW PACKAGE READY")
        print("="*70)
        
        # Analyze the distribution with explanations
        workflow_matches = self.generate_workflow_matches(matches_df, explanations)
        
        priority_counts = {}
        confidence_counts = {}
        quality_counts = {}
        
        for match in workflow_matches:
            priority_counts[match.review_priority] = priority_counts.get(match.review_priority, 0) + 1
            confidence_counts[match.confidence_level] = confidence_counts.get(match.confidence_level, 0) + 1
            quality_counts[match.quality_grade] = quality_counts.get(match.quality_grade, 0) + 1
        
        print(f"\nMatch Distribution:")
        print(f"  Total Matches: {len(workflow_matches)}")
        print(f"  Requirements Covered: {len(set(m.requirement_id for m in workflow_matches))}")
        
        print(f"\nBy Review Priority:")
        for priority, count in sorted(priority_counts.items()):
            estimated_hours = count * self._estimate_review_hours(priority)
            print(f"  {priority:15}: {count:3d} matches ({estimated_hours:5.1f} hours)")
        
        print(f"\nBy Confidence Level:")
        for confidence, count in sorted(confidence_counts.items()):
            print(f"  {confidence:8}: {count:3d} matches ({count/len(workflow_matches)*100:5.1f}%)")
        
        print(f"\nBy Requirement Quality:")
        for quality, count in sorted(quality_counts.items()):
            print(f"  {quality:10}: {count:3d} matches ({count/len(workflow_matches)*100:5.1f}%)")
        
        print(f"\nFiles Created:")
        for purpose, filepath in files_created.items():
            print(f"  {purpose:20}: {filepath}")
        
        print(f"\nImmediate Actions for Team Lead:")
        print(f"1. Open '{files_created['review_workbook']}' in Excel for detailed review")
        print(f"2. Review '{files_created['explanation_report']}' for visual summary")
        print(f"3. Assign engineers using '{files_created['action_items']}'")
        print(f"4. Address POOR/CRITICAL quality requirements first")
        
        # Estimate total effort
        total_hours = sum(self._estimate_review_hours(m.review_priority) for m in workflow_matches)
        quality_improvement_candidates = len([m for m in workflow_matches if m.quality_grade in ['POOR', 'CRITICAL']])
        
        print(f"\nEstimated Total Review Effort: {total_hours:.1f} hours")
        print(f"Requirements needing quality improvement: {quality_improvement_candidates}")
        
        return matches_df


def main():
    """Run enhanced workflow matching with full explainability."""
    
    print("="*70)
    print("ENHANCED WORKFLOW MATCHER - With Full Explainability")
    print("="*70)
    
    # Create enhanced workflow matcher
    workflow_matcher = EnhancedWorkflowMatcher()
    
    # Run complete workflow with explanations
    results = workflow_matcher.run_enhanced_workflow_matching(
        requirements_file="requirements.csv",
        activities_file="activities.csv",
        gold_pairs_file="manual_matches.csv",
        min_sim=0.35,
        top_n=5,
        out_file="enhanced_workflow_matches"
    )
    
    print(f"\nEnhanced Workflow Matching completed!")
    print(f"Check 'enhanced_engineering_review' folder for all deliverables")


if __name__ == "__main__":
    main()