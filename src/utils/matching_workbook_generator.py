"""
Fixed Matching Workbook Generator - Based on working project knowledge version
Focused on explanation-enhanced matching results with streamlined executive summary
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import logging
import json
import sys  
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
# Import project utilities
from src.utils.file_utils import SafeFileHandler
from src.utils.path_resolver import SmartPathResolver

logger = logging.getLogger(__name__)


class MatchingWorkbookGenerator:
    """Generate comprehensive Excel workbook for engineering teams to review matches."""
    
    def __init__(self, repo_manager=None):
        self.workbook_path = None
        if repo_manager is None:
            raise ValueError("Repository manager is required")
        self.repo_manager = repo_manager

        # Initialize utilities
        self.file_handler = SafeFileHandler(self.repo_manager)
        self.path_resolver = SmartPathResolver(self.repo_manager)
        
    def create_workbook(self, 
                       enhanced_df: pd.DataFrame,
                       evaluation_results: Optional[Dict] = None,
                       output_path: Optional[str] = None,
                       repo_manager=None) -> str:
        """Create comprehensive Excel workbook for engineering review."""
        
        # Setup repository manager
        if repo_manager is None:
            from src.utils.repository_setup import RepositoryStructureManager
            repo_manager = RepositoryStructureManager("outputs")
            repo_manager.setup_repository_structure()
        
        # Use proper output path
        if output_path is None:
            output_path = repo_manager.structure['engineering_review'] / "matching_workbook.xlsx"
        else:
            output_path = Path(output_path)
        
        # Data validation
        if enhanced_df is None or enhanced_df.empty:
            logger.error("❌ No data provided to workbook generator")
            raise ValueError("Enhanced DataFrame is empty or None")
        
        logger.info(f"📊 Creating workbook with {len(enhanced_df)} matches")
        
        # Create Excel writer
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # Tab 1: Executive Summary (README-style)
                try:
                    exec_summary = self._create_executive_summary(enhanced_df, evaluation_results)
                    exec_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
                    logger.info("✓ Created Executive Summary tab")
                except Exception as e:
                    logger.warning(f"Executive summary failed: {e}")
                    # Fallback summary
                    fallback_summary = pd.DataFrame([
                        {"Section": "Overview", "Metric": "Total Matches", "Value": len(enhanced_df), "Details": "Generated with fallback"}
                    ])
                    fallback_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
                
                # Tab 2-4: Confidence-based tabs (working version)
                try:
                    self._create_confidence_tabs(enhanced_df, writer)
                    logger.info("✓ Created confidence-based tabs")
                except Exception as e:
                    logger.error(f"Confidence tabs failed: {e}")
                    # Emergency fallback tab
                    enhanced_df.to_excel(writer, sheet_name='All Matches', index=False)
                
                # Tab 5: Detailed Explanations
                try:
                    self._create_explanations_tab(enhanced_df, writer)
                    logger.info("✓ Created Explanations tab")
                except Exception as e:
                    logger.warning(f"Explanations tab failed: {e}")
                
                # Format all worksheets
                try:
                    self._format_workbook(writer)
                except Exception as e:
                    logger.warning(f"Formatting failed: {e}")
        
        except Exception as e:
            logger.error(f"Excel creation failed: {e}")
            # CSV fallback
            fallback_path = output_path.with_suffix('.csv')
            enhanced_df.to_csv(fallback_path, index=False)
            return str(fallback_path)
        
        logger.info(f"✅ Matching workbook created: {output_path}")
        self.workbook_path = str(output_path)
        return str(output_path)
    
    def _create_executive_summary(self, enhanced_df: pd.DataFrame, 
                                evaluation_results: Optional[Dict]) -> pd.DataFrame:
        """Create technical README explaining the matching algorithm - aligned with actual implementation."""
        
        summary_data = []
        
        # === HEADER ===
        summary_data.extend([
            {'Section': 'REQUIREMENTS-TO-ACTIVITIES MATCHING ALGORITHM', 'Metric': '', 'Value': '', 'Details': f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # === DATASET OVERVIEW ===
        total_matches = len(enhanced_df)
        unique_reqs = enhanced_df['Requirement_ID'].nunique() if 'Requirement_ID' in enhanced_df.columns else 0
        unique_activities = enhanced_df['Activity_Name'].nunique() if 'Activity_Name' in enhanced_df.columns else 0
        
        summary_data.extend([
            {'Section': 'DATASET OVERVIEW', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Requirements Analyzed', 'Value': unique_reqs, 'Details': 'Unique requirements from input'},
            {'Section': '', 'Metric': 'Activities Available', 'Value': unique_activities, 'Details': 'Unique V&V activities in library'},
            {'Section': '', 'Metric': 'Matches Generated', 'Value': total_matches, 'Details': 'Requirement-activity pairs above threshold'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # === HOW MATCHING WORKS - Based on actual matcher.py ===
        summary_data.extend([
            {'Section': 'MATCHING ALGORITHM OVERVIEW', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Approach', 'Value': 'Multi-algorithm ensemble', 'Details': 'Combines four complementary matching techniques'},
            {'Section': '', 'Metric': 'Philosophy', 'Value': 'No single point of failure', 'Details': 'Each algorithm catches different types of valid matches'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # === ALGORITHM COMPONENTS - Aligned with matcher.py ===
        summary_data.extend([
            {'Section': 'ALGORITHM COMPONENTS', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': '1. SEMANTIC MATCHING', 'Value': 'AI-based understanding', 'Details': ''},
            {'Section': '', 'Metric': 'What it does', 'Value': 'Conceptual similarity', 'Details': 'Understands meaning beyond exact words'},
            {'Section': '', 'Metric': 'How it works', 'Value': 'Sentence transformers', 'Details': 'Uses all-MiniLM-L6-v2 model for 384-dim embeddings'},
            {'Section': '', 'Metric': 'Best for', 'Value': 'Paraphrased requirements', 'Details': '"Verify thrust" matches "Validate propulsion performance"'},
            {'Section': '', 'Metric': 'Limitations', 'Value': 'Context-free', 'Details': 'Cannot understand references like "previous requirement"'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': '2. BM25 (KEYWORD MATCHING)', 'Value': 'Traditional search', 'Details': ''},
            {'Section': '', 'Metric': 'What it does', 'Value': 'Exact term overlap', 'Details': 'Finds shared technical terminology'},
            {'Section': '', 'Metric': 'How it works', 'Value': 'TF-IDF variant', 'Details': 'Weights rare terms higher, common words lower'},
            {'Section': '', 'Metric': 'Best for', 'Value': 'Technical precision', 'Details': 'Excellent for acronyms, part numbers, specific terms'},
            {'Section': '', 'Metric': 'Special handling', 'Value': 'Short activities boost', 'Details': 'Extra weight for exact matches in brief activity names'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': '3. DOMAIN KNOWLEDGE', 'Value': 'Aerospace terminology', 'Details': ''},
            {'Section': '', 'Metric': 'What it does', 'Value': 'Recognizes critical terms', 'Details': 'Boosts matches with domain-specific vocabulary'},
            {'Section': '', 'Metric': 'Vocabulary includes', 'Value': 'V&V, safety, systems', 'Details': 'FMEA, FTA, qualification, verification, validation, etc.'},
            {'Section': '', 'Metric': 'Best for', 'Value': 'Industry compliance', 'Details': 'Ensures aerospace-relevant matches score higher'},
            {'Section': '', 'Metric': 'Coverage', 'Value': '~500 aerospace terms', 'Details': 'Curated from standards and domain expertise'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': '4. QUERY EXPANSION', 'Value': 'Synonym matching', 'Details': ''},
            {'Section': '', 'Metric': 'What it does', 'Value': 'Handles variations', 'Details': 'Matches equivalent terms and abbreviations'},
            {'Section': '', 'Metric': 'Examples', 'Value': 'S/C ↔ spacecraft', 'Details': 'verify ↔ validate, GN&C ↔ guidance navigation control'},
            {'Section': '', 'Metric': 'Best for', 'Value': 'Terminology variations', 'Details': 'Catches matches despite inconsistent wording'},
            {'Section': '', 'Metric': 'Limitations', 'Value': 'Expansion noise', 'Details': 'Can create false positives through over-expansion'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # === SCORE INTERPRETATION ===
        summary_data.extend([
            {'Section': 'COMBINED SCORE INTERPRETATION', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Score Range', 'Value': '0.0 to 1.0', 'Details': 'Weighted combination of all four algorithms'},
            {'Section': '', 'Metric': 'High (≥0.8)', 'Value': 'Strong alignment', 'Details': 'Multiple algorithms agree this is a good match'},
            {'Section': '', 'Metric': 'Medium (0.5-0.8)', 'Value': 'Probable match', 'Details': 'Some algorithms agree, needs engineering review'},
            {'Section': '', 'Metric': 'Low (<0.5)', 'Value': 'Weak alignment', 'Details': 'Limited algorithmic support, likely false positive'},
            {'Section': '', 'Metric': 'Filtering', 'Value': 'Min 0.35, Top 5', 'Details': 'Keeps only scores ≥0.35, max 5 matches per requirement'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # === SCORE BREAKDOWN PATTERNS ===
        summary_data.extend([
            {'Section': 'READING SCORE BREAKDOWNS', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Format', 'Value': 'Sem:X | BM25:Y | Dom:Z | QE:W', 'Details': 'Individual algorithm contributions'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'RELIABLE PATTERNS', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Balanced scores', 'Value': 'Multiple >0.5', 'Details': 'Algorithms agree from different perspectives'},
            {'Section': '', 'Metric': 'High Semantic + BM25', 'Value': 'Concept + terms align', 'Details': 'Both meaning and terminology match'},
            {'Section': '', 'Metric': 'High BM25 + Domain', 'Value': 'Technical match', 'Details': 'Precise aerospace terminology alignment'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'SUSPICIOUS PATTERNS', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Only Semantic high', 'Value': 'Conceptual only', 'Details': 'May understand concept but wrong activity type'},
            {'Section': '', 'Metric': 'Only Query Expansion high', 'Value': 'Synonym artifact', 'Details': 'Matched through expansion, not core content'},
            {'Section': '', 'Metric': 'Single component match', 'Value': 'Narrow evidence', 'Details': 'Other algorithms found no support'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # === VALIDATION PERFORMANCE ===
        if evaluation_results and 'metrics' in evaluation_results:
            metrics = evaluation_results['metrics']
            summary_data.extend([
                {'Section': 'ALGORITHM VALIDATION', 'Metric': '', 'Value': '', 'Details': ''},
                {'Section': '', 'Metric': 'Benchmark', 'Value': 'Expert manual traces', 'Details': 'Tested against your existing requirement-activity mappings'},
                {'Section': '', 'Metric': 'F1 Score @ Top 5', 'Value': f"{metrics.get('f1_at_5', 0):.1%}", 'Details': 'Overall accuracy (balance of precision and recall)'},
                {'Section': '', 'Metric': 'Coverage', 'Value': f"{metrics.get('coverage', 0):.1%}", 'Details': 'Requirements with at least one valid match found'},
                {'Section': '', 'Metric': 'Perfect Matches', 'Value': f"{metrics.get('perfect_matches', 0)}/{metrics.get('total_evaluated', 0)}", 'Details': "Algorithm's top pick matched expert's top pick"},
                {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
            ])
        
        # === MATCH QUALITY LEVELS ===
        summary_data.extend([
            {'Section': 'MATCH QUALITY LEVELS', 'Metric': '', 'Value': '', 'Details': 'Based on combined score thresholds'},
            {'Section': '', 'Metric': 'EXCELLENT', 'Value': 'Score ≥ 0.6', 'Details': 'Very strong match, high confidence'},
            {'Section': '', 'Metric': 'GOOD', 'Value': 'Score 0.45-0.6', 'Details': 'Good match, moderate confidence'},
            {'Section': '', 'Metric': 'MODERATE', 'Value': 'Score 0.3-0.45', 'Details': 'Possible match, needs review'},
            {'Section': '', 'Metric': 'WEAK', 'Value': 'Score < 0.3', 'Details': 'Poor match, likely false positive'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # === KEY LIMITATIONS ===
        summary_data.extend([
            {'Section': 'KNOWN LIMITATIONS', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Context understanding', 'Value': 'Not supported', 'Details': 'Cannot resolve "it", "previous", "aforementioned"'},
            {'Section': '', 'Metric': 'Project-specific terms', 'Value': 'May miss', 'Details': 'Custom acronyms or unique project terminology'},
            {'Section': '', 'Metric': 'Very short text', 'Value': 'Less reliable', 'Details': 'Requirements <5 words have less signal'},
            {'Section': '', 'Metric': 'Negative requirements', 'Value': 'Challenging', 'Details': '"Shall not" semantics sometimes misunderstood'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        return pd.DataFrame(summary_data)    
    
    def _create_confidence_tabs(self, enhanced_df: pd.DataFrame, writer):
        """Create confidence-based tabs - fixed to use correct column names."""
        
        # Debug: Print available columns
        logger.info(f"Available columns: {list(enhanced_df.columns)}")
        
        # Find score column - try both formats
        score_col = None
        for candidate in ['Combined Score', 'Combined_Score', 'combined_score', 'score']:
            if candidate in enhanced_df.columns:
                score_col = candidate
                logger.info(f"Found score column: {score_col}")
                break
        
        if score_col is None:
            logger.error(f"No score column found! Available columns: {list(enhanced_df.columns)}")
            enhanced_df.to_excel(writer, sheet_name='All Matches', index=False)
            return
        
        # Load explanations if available
        explanations_data = self._load_explanations()
        logger.info(f"Loaded {len(explanations_data)} explanations")
        
        # Define confidence levels
        confidence_levels = [
            ("High Confidence (≥0.8)", enhanced_df[enhanced_df[score_col] >= 0.8]),
            ("Medium Confidence (0.5-0.8)", enhanced_df[(enhanced_df[score_col] >= 0.5) & 
                                                        (enhanced_df[score_col] < 0.8)]),
            ("Low Confidence (<0.5)", enhanced_df[enhanced_df[score_col] < 0.5])
        ]
        
        tabs_created = 0
        
        for tab_name, df_subset in confidence_levels:
            if len(df_subset) > 0:
                try:
                    logger.info(f"Creating {tab_name} with {len(df_subset)} matches")
                    
                    # Select columns for engineering review
                    review_columns = self._get_review_columns(df_subset)
                    logger.info(f"Using columns: {review_columns}")
                    review_df = df_subset[review_columns].copy()
                    
                    # Sort by score
                    review_df = review_df.sort_values(score_col, ascending=False)
                    
                    # Add explanation columns (this creates Score Breakdown column)
                    try:
                        review_df['Match Explanation'] = review_df.apply(
                            lambda row: self._get_explanation_text(row, explanations_data), axis=1
                        )
                        review_df['Key Evidence'] = review_df.apply(
                            lambda row: self._get_key_evidence(row, explanations_data), axis=1
                        )
                        review_df['Score Breakdown'] = review_df.apply(
                            lambda row: self._get_score_breakdown(row, explanations_data), axis=1
                        )
                        logger.info("✓ Added explanation columns")
                    except Exception as e:
                        logger.warning(f"Could not add explanations: {e}")
                    
                    # Add review column
                    review_df['Engineering Notes'] = ''
                    
                    # Write to Excel
                    review_df.to_excel(writer, sheet_name=tab_name, index=False)
                    tabs_created += 1
                    
                    logger.info(f"✓ Created {tab_name} tab: {len(review_df)} matches")
                    
                except Exception as e:
                    logger.error(f"Failed to create {tab_name}: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                logger.info(f"Skipping {tab_name} - no matches in this range")
        
        # Ensure at least one data tab exists
        if tabs_created == 0:
            logger.warning("No confidence tabs created, creating fallback All Matches tab")
            enhanced_df.to_excel(writer, sheet_name='All Matches', index=False)

    def _get_review_columns(self, df: pd.DataFrame) -> List[str]:
        """Get review columns - try both underscore and space formats."""
        
        base_columns = []
        
        # Find ID column - try both formats
        for id_col in ['ID', 'Requirement_ID', 'requirement_id']:
            if id_col in df.columns:
                base_columns.append(id_col)
                logger.info(f"Found ID column: {id_col}")
                break
        
        # Find name column - try both formats
        for name_col in ['Requirement Name', 'Requirement_Name', 'requirement_name']:
            if name_col in df.columns:
                base_columns.append(name_col)
                logger.info(f"Found name column: {name_col}")
                break
        
        # Find text column - try both formats
        for text_col in ['Requirement Text', 'Requirement_Text', 'requirement_text']:
            if text_col in df.columns:
                base_columns.append(text_col)
                logger.info(f"Found text column: {text_col}")
                break
        
        # Find activity column - try both formats
        for act_col in ['Activity Name', 'Activity_Name', 'activity_name']:
            if act_col in df.columns:
                base_columns.append(act_col)
                logger.info(f"Found activity column: {act_col}")
                break
        
        # Find score column - try both formats
        for score_col in ['Combined Score', 'Combined_Score', 'combined_score']:
            if score_col in df.columns:
                base_columns.append(score_col)
                logger.info(f"Found score column: {score_col}")
                break
        
        # CRITICAL: Do NOT include individual score columns
        # This forces the Score Breakdown column to be used instead
        
        # Add quality information if available
        quality_columns = ['Quality_Grade', 'Quality_Score']
        for col in quality_columns:
            if col in df.columns:
                base_columns.append(col)
                logger.info(f"Found quality column: {col}")
        
        # Debug: show what we found
        logger.info(f"Selected review columns: {base_columns}")
        
        # Ensure we have at least some columns
        if len(base_columns) < 2:
            logger.warning("Too few columns found, using all columns as fallback")
            base_columns = list(df.columns)
        
        # Only return columns that exist in the DataFrame
        final_columns = [col for col in base_columns if col in df.columns]
        logger.info(f"Final columns to use: {final_columns}")
        return final_columns
    
    def _create_explanations_tab(self, enhanced_df: pd.DataFrame, writer):
        """Create detailed explanations tab showing algorithm reasoning."""
        
        explanations_data = self._load_explanations()
        if not explanations_data:
            logger.warning("No explanations data available")
            return
        
        # Convert explanations to DataFrame
        explanations_list = []
        for key, exp in explanations_data.items():
            req_id, activity_name = key
            scores_obj = exp.get('scores', {})
            explanations_list.append({
                'ID': req_id,                              # MATCHER OUTPUT COMPLIANT
                'Activity Name': activity_name,            # MATCHER OUTPUT COMPLIANT  
                'Combined Score': exp.get('combined_score', 0),        # MATCHER OUTPUT COMPLIANT
                'Dense Semantic': scores_obj.get('semantic', 0),       # MATCHER OUTPUT COMPLIANT
                'BM25 Score': scores_obj.get('bm25', 0),               # MATCHER OUTPUT COMPLIANT
                'Domain Weighted': scores_obj.get('domain', 0),        # MATCHER OUTPUT COMPLIANT
                'Query Expansion': scores_obj.get('query_expansion', 0), # MATCHER OUTPUT COMPLIANT
                'Shared_Terms': ', '.join(exp.get('shared_terms', [])),
                'Semantic_Explanation': exp.get('explanations', {}).get('semantic', ''),
                'BM25_Explanation': exp.get('explanations', {}).get('bm25', ''),
                'Domain_Explanation': exp.get('explanations', {}).get('domain', ''),
                'Query_Expansion_Explanation': exp.get('explanations', {}).get('query_expansion', ''),
                'Match_Quality': exp.get('match_quality', 'Unknown')
            })
        
        if explanations_list:
            explanations_df = pd.DataFrame(explanations_list)
            # Sort by Combined Score (MATCHER OUTPUT COMPLIANT)
            explanations_df = explanations_df.sort_values('Combined Score', ascending=False)
            explanations_df.to_excel(writer, sheet_name='Detailed Explanations', index=False)
            logger.info(f"✓ Created Detailed Explanations tab: {len(explanations_df)} explanations")
    
    def _load_explanations(self) -> Dict:
        """Load explanations from JSON file."""
        
        # Try multiple possible locations
        possible_paths = [
            self.repo_manager.structure['matching_results'] / "aerospace_matches_explanations.json",
            Path("outputs/matching_results/aerospace_matches_explanations.json"),
            Path("aerospace_matches_explanations.json")
        ]
        
        for explanations_file in possible_paths:
            if explanations_file.exists():
                try:
                    with open(explanations_file, 'r', encoding='utf-8') as f:
                        explanations = json.load(f)
                        # Create lookup dictionary - fix key mapping
                        explanations_dict = {}
                        for exp in explanations:
                            # Handle both possible JSON structures
                            req_id = exp.get('requirement_id') or exp.get('Requirement_ID')
                            act_name = exp.get('activity_name') or exp.get('Activity_Name')
                            
                            if req_id and act_name:
                                key = (str(req_id), str(act_name))
                                explanations_dict[key] = exp
                        
                        logger.info(f"✅ Loaded {len(explanations_dict)} match explanations")
                        return explanations_dict
                except Exception as e:
                    logger.warning(f"Could not load explanations from {explanations_file}: {e}")
        
        logger.warning("No explanations file found")
        return {}

    def _get_explanation_text(self, row, explanations_data: Dict) -> str:
        """Get explanation text using the original working approach."""
        
        # Use actual column names from your data (with underscores)
        if 'Requirement_ID' in row and 'Activity_Name' in row:
            key = (str(row['Requirement_ID']), str(row['Activity_Name']))
            if key in explanations_data:
                exp = explanations_data[key]
                
                # Use the ORIGINAL WORKING APPROACH from the document
                sem_level = exp.get('semantic_similarity_level', 'Unknown')
                parts = [f"{sem_level} semantic match"]
                scores = exp.get('scores', {})
                
                if scores.get('bm25', 0) > 0.5:
                    parts.append("strong term overlap")
                if scores.get('domain', 0) > 0.3:
                    parts.append("aerospace terms matched")
                    
                return "; ".join(parts)
        
        # Fallback based on available score columns (also from original)
        score = 0
        for score_col in ['Combined_Score', 'Combined_Score', 'combined_score', 'score']:
            if score_col in row and pd.notna(row[score_col]):
                score = row[score_col]
                break
        
        # Original fallback logic
        if score >= 0.8:
            return "Very strong match across multiple dimensions"
        elif score >= 0.6:
            return "Good match with solid evidence"
        elif score >= 0.4:
            return "Moderate match, review recommended"
        else:
            return "Weak match, manual verification needed"

    def _get_key_evidence(self, row, explanations_data: Dict) -> str:
        """Get key evidence (shared terms) for a match."""
        
        # Use actual column names from your data (with underscores)
        if 'Requirement_ID' in row and 'Activity_Name' in row:
            key = (str(row['Requirement_ID']), str(row['Activity_Name']))
            if key in explanations_data:
                exp = explanations_data[key]
                shared_terms = exp.get('shared_terms', [])
                if shared_terms:
                    return f"Key terms: {', '.join(shared_terms[:5])}"
                    
                # Fallback to any domain explanation
                domain_exp = exp.get('explanations', {}).get('domain', '')
                if domain_exp and domain_exp != 'N/A' and len(domain_exp) < 100:
                    return domain_exp
        
        return "Check individual scores"

    def _get_score_breakdown(self, row, explanations_data: Dict) -> str:
        """Get compact score breakdown from explanations or row data."""
        
        breakdown = []
        
        # Try to get scores from explanations first (nested structure)
        if 'Requirement_ID' in row and 'Activity_Name' in row:
            key = (str(row['Requirement_ID']), str(row['Activity_Name']))
            if key in explanations_data:
                exp = explanations_data[key]
                scores_obj = exp.get('scores', {})
                # Get scores from nested structure
                if scores_obj.get('semantic') is not None:
                    breakdown.append(f"Sem:{scores_obj['semantic']:.2f}")
                if scores_obj.get('bm25') is not None:
                    breakdown.append(f"BM25:{scores_obj['bm25']:.2f}")
                if scores_obj.get('domain') is not None:
                    breakdown.append(f"Dom:{scores_obj['domain']:.2f}")
                if scores_obj.get('query_expansion') is not None:
                    breakdown.append(f"QE:{scores_obj['query_expansion']:.2f}")
                    
                if breakdown:
                    return " | ".join(breakdown)
        
        # Fallback: Try to get scores from row data using actual column names (with underscores)
        score_mappings = [
            ('Semantic_Score', 'Sem'),     # Your actual column name
            ('BM25_Score', 'BM25'),        # Your actual column name
            ('Domain_Score', 'Dom'),       # Your actual column name  
            ('Query_Expansion_Score', 'QE') # Your actual column name
        ]
        
        for col, label in score_mappings:
            if col in row and pd.notna(row[col]):
                breakdown.append(f"{label}:{row[col]:.2f}")
        
        return " | ".join(breakdown) if breakdown else f"Combined: {row.get('Combined_Score', 'N/A')}"

    def _format_workbook(self, writer):
        """Apply enhanced formatting with special handling for different sheets."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
            
            workbook = writer.book
            
            # Define formatting styles
            header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Format each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                try:
                    # Special handling for Executive Summary - NO TABLE, just formatting
                    if sheet_name == 'Executive Summary':
                        # Format headers
                        if worksheet.max_row > 0:
                            for cell in worksheet[1]:
                                if cell.value:
                                    cell.font = Font(bold=True, size=12)
                                    cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                                    cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                        # Auto-fit columns for Executive Summary
                        for column_cells in worksheet.columns:
                            column_letter = get_column_letter(column_cells[0].column)
                            
                            # Calculate max width
                            max_length = 0
                            for cell in column_cells:
                                try:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                except:
                                    pass
                            
                            # Set width with reasonable limits
                            adjusted_width = min(max_length + 2, 60)
                            worksheet.column_dimensions[column_letter].width = max(adjusted_width, 15)
                        
                        # Add some styling to section headers
                        for row_idx in range(2, worksheet.max_row + 1):
                            section_cell = worksheet.cell(row_idx, 1)
                            if section_cell.value and section_cell.value.strip() and worksheet.cell(row_idx, 2).value == '':
                                # This is a section header
                                for col_idx in range(1, 5):
                                    cell = worksheet.cell(row_idx, col_idx)
                                    cell.font = Font(bold=True, size=11)
                                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        
                        logger.info("✓ Formatted Executive Summary (no table)")
                    
                    # Create Excel tables for confidence tabs and explanations
                    elif sheet_name in ['High Confidence (≥0.8)', 'Medium Confidence (0.5-0.8)', 
                                    'Low Confidence (<0.5)', 'Detailed Explanations']:
                        # Determine the table range
                        max_row = worksheet.max_row
                        max_col = worksheet.max_column
                        
                        if max_row > 1 and max_col > 0:  # Ensure there's data beyond headers
                            # Create table reference using column letters
                            table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
                            
                            # Clean table name (remove special characters)
                            table_name = sheet_name.replace(' ', '_').replace('(', '').replace(')', '').replace('≥', 'GTE').replace('<', 'LT').replace('-', '_').replace('.', '_')
                            
                            # Create and add table
                            tab = Table(displayName=table_name, ref=table_ref)
                            style = TableStyleInfo(
                                name="TableStyleMedium9", 
                                showFirstColumn=False,
                                showLastColumn=False, 
                                showRowStripes=True, 
                                showColumnStripes=False
                            )
                            tab.tableStyleInfo = style
                            worksheet.add_table(tab)
                            logger.info(f"✓ Created table for {sheet_name}")
                    
                    # Apply column formatting for all sheets (except Executive Summary which was already done)
                    if sheet_name != 'Executive Summary':
                        # Column width calculation and formatting
                        for column_cells in worksheet.columns:
                            column_letter = get_column_letter(column_cells[0].column)
                            
                            # Skip empty columns
                            if not any(cell.value for cell in column_cells):
                                continue
                            
                            # Get column header name
                            header_value = column_cells[0].value
                            column_name = str(header_value) if header_value else ""
                            
                            # Calculate max width based on content
                            max_length = 0
                            for cell in column_cells:
                                try:
                                    if cell.value:
                                        # Consider line breaks in the content
                                        cell_value = str(cell.value)
                                        lines = cell_value.split('\n')
                                        max_line_length = max(len(line) for line in lines) if lines else len(cell_value)
                                        max_length = max(max_length, max_line_length)
                                except:
                                    pass
                            
                            # Adjust width with a multiplier for character width
                            adjusted_width = min(max_length * 1.1 + 2, 100)  # Cap at 100
                            
                            # Apply column-specific constraints and formatting
                            if 'Requirement_Text' in column_name:
                                # Requirement Text: Set reasonable width with wrapping
                                worksheet.column_dimensions[column_letter].width = max(40, min(adjusted_width, 80))
                                
                                # Apply wrap text to all data cells in this column
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:  # Skip header row
                                        cell.alignment = Alignment(
                                            wrap_text=True, 
                                            vertical='top', 
                                            horizontal='left'
                                        )
                                logger.info(f"✓ Auto-fit {column_name} with text wrapping")
                            
                            elif any(keyword in column_name for keyword in ['_ID', 'ID']):
                                # ID columns: Narrow width
                                worksheet.column_dimensions[column_letter].width = max(10, min(adjusted_width, 20))
                                
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                logger.info(f"✓ Auto-fit {column_name} (ID column)")
                            
                            elif 'Score' in column_name and 'Breakdown' not in column_name:
                                # Score columns: Narrow width, centered
                                worksheet.column_dimensions[column_letter].width = max(12, min(adjusted_width, 18))
                                
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                        # Format scores to 3 decimal places
                                        try:
                                            if isinstance(cell.value, (int, float)):
                                                cell.number_format = '0.000'
                                        except:
                                            pass
                                logger.info(f"✓ Auto-fit {column_name} (Score column)")
                            
                            elif any(keyword in column_name for keyword in ['Explanation', 'Evidence', 'Notes', 'Breakdown']):
                                # Explanation columns: Medium width
                                worksheet.column_dimensions[column_letter].width = max(25, min(adjusted_width, 100))
                                
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:
                                        cell.alignment = Alignment(
                                            horizontal='left', 
                                            vertical='top', 
                                            wrap_text=False
                                        )
                                logger.info(f"✓ Auto-fit {column_name} (Explanation column)")
                            
                            elif any(keyword in column_name for keyword in ['Name', 'Activity']):
                                # Name and Activity columns: Medium width
                                worksheet.column_dimensions[column_letter].width = max(20, min(adjusted_width, 45))
                                
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:
                                        cell.alignment = Alignment(horizontal='left', vertical='center')
                                logger.info(f"✓ Auto-fit {column_name} (Name/Activity column)")
                            
                            else:
                                # Default columns: Auto-fit with reasonable constraints
                                worksheet.column_dimensions[column_letter].width = max(12, min(adjusted_width, 50))
                                
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:
                                        cell.alignment = Alignment(horizontal='left', vertical='center')
                                logger.info(f"✓ Auto-fit {column_name} (Default column)")
                        
                        # Set intelligent row heights (not for Executive Summary)
                        for row_idx in range(2, worksheet.max_row + 1):  # Skip header
                            # Check if this row has requirement text that might wrap
                            has_wrapping_text = False
                            max_text_length = 0
                            
                            for cell in worksheet[row_idx]:
                                if cell.value:
                                    cell_text = str(cell.value)
                                    # Get column header to check if it's Requirement_Text
                                    header_cell = worksheet.cell(1, cell.column)
                                    column_name = str(header_cell.value) if header_cell.value else ""
                                    
                                    if 'Requirement_Text' in column_name and len(cell_text) > 50:
                                        has_wrapping_text = True
                                        max_text_length = max(max_text_length, len(cell_text))
                            
                            # Set row height based on content
                            if has_wrapping_text:
                                # Calculate height based on text length and column width
                                # Assuming ~80 characters per line with 80-width column
                                estimated_lines = max(2, min(8, max_text_length // 80))
                                worksheet.row_dimensions[row_idx].height = 15 + (estimated_lines * 15)
                            else:
                                # Standard height for other rows
                                worksheet.row_dimensions[row_idx].height = 20
                    
                    # Set header row height (all sheets)
                    if worksheet.max_row >= 1:
                        worksheet.row_dimensions[1].height = 25
                    
                    # Freeze panes (header row) - all sheets except Executive Summary
                    if sheet_name != 'Executive Summary':
                        worksheet.freeze_panes = worksheet['A2']
                    
                    logger.info(f"✓ Formatted sheet: {sheet_name}")
                
                except Exception as e:
                    logger.warning(f"Could not format sheet {sheet_name}: {e}")
                    import traceback
                    traceback.print_exc()
        
        except ImportError:
            logger.warning("openpyxl styling not available")
        except Exception as e:
            logger.warning(f"Workbook formatting failed: {e}")
            import traceback
            traceback.print_exc()

def create_matching_workbook(enhanced_df: pd.DataFrame, 
                           evaluation_results: Optional[Dict] = None,
                           output_path: str = "outputs/engineering_review/matching_workbook.xlsx",
                           repo_manager=None) -> str:
    """
    Convenience function to create streamlined matching workbook.
    
    Args:
        enhanced_df: Enhanced predictions DataFrame
        evaluation_results: Optional evaluation results
        output_path: Where to save the workbook
        repo_manager: Repository manager instance
        
    Returns:
        Path to created workbook
    """
    generator = MatchingWorkbookGenerator(repo_manager)
    return generator.create_workbook(enhanced_df, evaluation_results, output_path, repo_manager)