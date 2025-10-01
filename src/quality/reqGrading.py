"""
Enhanced Requirements Quality Analyzer with 4-Tab Excel Output
Provides comprehensive analysis with Summary, Quality, INCOSE, and Semantic tabs.
"""

import pandas as pd
import numpy as np
from typing import List, Dict, Tuple, Optional, Any, Set
import logging
from pathlib import Path
import spacy
import json
import re
from dataclasses import dataclass, field, asdict
from collections import Counter, defaultdict
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Import utilities
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from src.utils.file_utils import SafeFileHandler
from src.utils.path_resolver import SmartPathResolver
from src.utils.repository_setup import RepositoryStructureManager

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class QualityMetrics:
    """Container for requirement quality metrics."""
    quality_score: float
    clarity_score: float
    completeness_score: float
    verifiability_score: float
    atomicity_score: float
    consistency_score: float
    incose_compliance_score: float
    semantic_quality_score: float
    total_issues: int
    severity_breakdown: Dict[str, int]

@dataclass
class INCOSEAnalysis:
    """INCOSE pattern analysis results."""
    best_pattern: str
    compliance_score: float
    components_found: Dict[str, Optional[str]]
    missing_required: List[str]
    missing_optional: List[str]
    suggestions: List[str]
    template_recommendation: str

@dataclass
class SemanticAnalysis:
    """Semantic analysis results."""
    similarity_issues: List[Dict]
    contextual_ambiguities: List[str]
    entity_completeness: Dict[str, List[str]]
    tone_issues: List[str]
    improvement_suggestions: List[str]

class INCOSEPatternAnalyzer:
    """INCOSE requirements pattern analyzer."""
    
    def __init__(self, nlp):
        self.nlp = nlp
        self.patterns = {
            'functional_performance': {
                'name': 'Functional/Performance',
                'required': ['AGENT', 'FUNCTION', 'PERFORMANCE'],
                'optional': ['INTERFACE_OUTPUT', 'TIMING', 'EVENT_TRIGGER', 'INTERFACE_INPUT', 'CONDITION'],
                'template': "The {AGENT} shall {FUNCTION} in accordance with {INTERFACE_OUTPUT} with {PERFORMANCE} [and {TIMING} upon {EVENT_TRIGGER} in accordance with {INTERFACE_INPUT}] while in {CONDITION}",
                'description': "Specifies what the system shall do and how well"
            },
            'suitability': {
                'name': 'Suitability',
                'required': ['AGENT', 'CHARACTERISTIC', 'PERFORMANCE'],
                'optional': ['CONDITION', 'CONDITION_DURATION'],
                'template': "The {AGENT} shall exhibit {CHARACTERISTIC} with {PERFORMANCE} while {CONDITION} [for {CONDITION_DURATION}]",
                'description': "Specifies quality characteristics the system must exhibit"
            },
            'environments': {
                'name': 'Environmental',
                'required': ['AGENT', 'CHARACTERISTIC', 'ENVIRONMENT'],
                'optional': ['EXPOSURE_DURATION'],
                'template': "The {AGENT} shall exhibit {CHARACTERISTIC} during/after exposure to {ENVIRONMENT} [for {EXPOSURE_DURATION}]",
                'description': "Specifies behavior under environmental conditions"
            },
            'design': {
                'name': 'Design Constraint',
                'required': ['AGENT', 'DESIGN_CONSTRAINTS'],
                'optional': ['PERFORMANCE', 'CONDITION'],
                'template': "The {AGENT} shall exhibit {DESIGN_CONSTRAINTS} [in accordance with {PERFORMANCE} while in {CONDITION}]",
                'description': "Specifies design limitations or constraints"
            }
        }
        
        # Component extraction patterns

        self.component_patterns = {
            'PERFORMANCE': [
                # ========== PERCENTAGE PATTERNS (Most Specific First) ==========
                # Percentage with comparative operators
                r'(?:less\s+than|fewer\s+than|below|under)\s+\d+(?:\.\d+)?\s*(?:%|percent)\s+(?:variation|degradation|error|deviation)\b',
                r'(?:at\s+least|minimum\s+of|no\s+less\s+than|greater\s+than|more\s+than|exceeds?|above)\s+\d+(?:\.\d+)?\s*(?:%|percent)\b',
                r'(?:at\s+most|maximum\s+of|no\s+more\s+than|less\s+than|fewer\s+than|below|under)\s+\d+(?:\.\d+)?\s*(?:%|percent)\b',
                r'(?:equal\s+to|exactly|precisely)\s+\d+(?:\.\d+)?\s*(?:%|percent)\b',
                r'(?:between|from|ranging\s+from)\s+\d+(?:\.\d+)?\s+(?:and|to)\s+\d+(?:\.\d+)?\s*(?:%|percent)\b',
                
                # Percentage with tolerance (±)
                r'(?:within\s+)?(?:±|plus\s+or\s+minus)\s*\d+(?:\.\d+)?\s*(?:%|percent)\b',
                
                # Standalone percentage
                r'\d+(?:\.\d+)?\s*(?:%|percent)\s+(?:of|or\s+(?:more|less|higher|lower|greater))\b',
                r'\d+(?:\.\d+)?\s*%\s+(?:accuracy|precision|efficiency|effectiveness|reliability|availability)\b',
                
                # ========== TEMPERATURE PATTERNS (PERFORMANCE levels) ==========
                # Temperature with tolerance (±)
                r'(?:within\s+)?(?:±|plus\s+or\s+minus)\s*\d+(?:\.\d+)?\s*(?:°C|°F|K|celsius|fahrenheit|kelvin|degrees?)\b',
                r'(?:±|plus\s+or\s+minus)\s*\d+(?:\.\d+)?\s*(?:°C|°F|K|degrees?)\s+(?:of|from|around)\b',
                
                # Temperature ranges
                r'temperature\s+(?:range\s+)?(?:of\s+|from\s+|between\s+)?(?:-?\d+(?:\.\d+)?)\s*(?:to\s+(?:-?\d+(?:\.\d+)?)\s*)?(?:°C|°F|K|celsius|fahrenheit|kelvin)\b',
                r'(?:thermal\s+)?(?:cycling|variation|gradient|shock)\s+(?:of\s+|from\s+|between\s+)?\d+(?:\.\d+)?\s*(?:°C|°F|K)\b',
                
                # Temperature levels (specific values)
                r'temperature\s+(?:of\s+|at\s+)?(?:-?\d+(?:\.\d+)?)\s*(?:°C|°F|K)\b',
                r'(?:at|within)\s+(?:-?\d+(?:\.\d+)?)\s*(?:to\s+(?:-?\d+(?:\.\d+)?)\s*)?(?:°C|°F|K)\b',
                
                # ========== TIME PATTERNS WITH UNITS ==========
                r'within\s+\d+(?:\.\d+)?\s*(?:ms|milliseconds?|μs|microseconds?|ns|nanoseconds?|sec|seconds?|min|minutes?|hr|hours?|days?|weeks?|months?|years?)\b',
                r'(?:within\s+)?(?:±|plus\s+or\s+minus)\s*\d+\s+days?\b',
                r'(?:after|following|before|prior\s+to)\s+\d+(?:\.\d+)?\s*(?:ms|milliseconds?|sec|seconds?|min|minutes?|hr|hours?|days?|weeks?)\b',
                
                # ========== TOLERANCE PATTERNS WITH UNITS ==========
                # General tolerance with physical units
                r'(?:within\s+)?(?:±|plus\s+or\s+minus)\s*\d+(?:\.\d+)?\s*(?:V|A|W|kg|g|m|cm|mm|km|ft|in|Pa|psi|Hz|rpm|N|lbf)\b',
                r'(?:±|plus\s+or\s+minus|tolerance\s+of)\s*\d+(?:\.\d+)?\s*(?:%|percent|ppm|parts?\s+per\s+million)\b',
                r'up\s+to\s+\d+(?:\.\d+)?\s*(?:km|meters?|m|kg|g|W|V|A|Hz|years?|months?|days?|hours?)\b',
                r'(?:accuracy|precision|resolution|tolerance|error)\s+(?:of\s+)?(?:±|less\s+than\s+|better\s+than\s+)?\d+(?:\.\d+)?\s*(?:%|ppm|bits?|degrees?|°|meters?|m|km|feet?|ft|inches?|in)\b',
                
                # ========== SPECIFIC UNIT PATTERNS ==========
                # Aerospace thrust/force
                r'(?:thrust|force)\s+(?:of\s+)?(?:up\s+to\s+|at\s+least\s+|minimum\s+|maximum\s+)?\d+(?:\.\d+)?\s*(?:N|newtons?|lbf|pounds?\s+force|kN|kilonewtons?)\b',
                
                # Velocity/speed
                r'(?:velocity|speed)\s+(?:of\s+)?(?:up\s+to\s+|at\s+least\s+)?\d+(?:\.\d+)?\s*(?:m/s|meters?\s+per\s+second|km/h|mph|ft/s|knots?)\b',
                
                # Power/energy
                r'(?:power|consumption|energy)\s+(?:of\s+)?(?:less\s+than\s+|maximum\s+|minimum\s+)?\d+(?:\.\d+)?\s*(?:W|watts?|kW|kilowatts?|mW|milliwatts?|kWh|Wh)\b',
                
                # Electrical
                r'(?:voltage|current)\s+(?:of\s+)?(?:±|tolerance\s+)?\d+(?:\.\d+)?\s*(?:V|volts?|A|amps?|amperes?|mA|milliamps?|kV|mV)\b',
                
                # Mechanical environment (PERFORMANCE levels - "how well")
                r'(?:vibration|acceleration|shock|impact)\s+(?:of\s+|up\s+to\s+)?(?:\d+(?:\.\d+)?(?:[eE][+-]?\d+)?|\d+\^\-?\d+)\s*(?:g|G|m/s²|Hz|grms)\b',
                r'(?:mechanical\s+)?(?:stress|strain|load|loading|force)\s+(?:of\s+|up\s+to\s+)?\d+(?:\.\d+)?\s*(?:Pa|psi|N|lbf|MPa|kPa)\b',
                
                # Pressure levels (PERFORMANCE - measurable criteria)
                r'(?:pressure|vacuum)\s+(?:of\s+|range\s+|from\s+|between\s+)?(?:\d+(?:\.\d+)?\s*(?:to\s+\d+(?:\.\d+)?\s*)?)?(?:Pa|psi|torr|atm|bar|mbar|kPa|MPa)\b',
                
                # Humidity levels (PERFORMANCE - specific values)
                r'(?:humidity|moisture|condensation)\s+(?:of\s+|from\s+|between\s+)?\d+(?:\.\d+)?\s*(?:to\s+\d+(?:\.\d+)?\s*)?(?:%|percent|RH)\b',
                
                # Altitude levels (PERFORMANCE - specific values)
                r'(?:altitude|elevation)\s+(?:of\s+|from\s+|between\s+|up\s+to\s+)?\d+(?:\.\d+)?\s*(?:to\s+\d+(?:\.\d+)?\s*)?(?:m|meters?|km|ft|feet)\b',
                
                # Radiation dose levels (PERFORMANCE - measurable criteria)
                r'(?:total\s+ionizing\s+dose|TID|total\s+dose)\s+(?:of\s+|up\s+to\s+)?\d+(?:\.\d+)?\s*(?:rad|krad|Mrad|Gy|kGy)\b',
                
                # Frequency/rate
                r'(?:frequency|rate|bandwidth)\s+(?:of\s+)?(?:up\s+to\s+|at\s+least\s+)?\d+(?:\.\d+)?\s*(?:Hz|hertz|kHz|MHz|GHz|rpm|revolutions?\s+per\s+minute|cycles?\s+per\s+second)\b',
                
                # Mass/weight
                r'(?:mass|weight)\s+(?:of\s+)?(?:less\s+than\s+|maximum\s+|minimum\s+)?\d+(?:\.\d+)?\s*(?:kg|kilograms?|g|grams?|lb|pounds?|oz|ounces?|tons?|mg|milligrams?)\b',
                
                # Distance/length/range
                r'(?:distance|length|range)\s+(?:of\s+)?(?:up\s+to\s+|at\s+least\s+|minimum\s+|maximum\s+)?\d+(?:\.\d+)?\s*(?:m|meters?|km|kilometers?|ft|feet|miles?|nautical\s+miles?|mm|cm|in|inches?)\b',
                
                # Data rate/throughput/bandwidth
                r'(?:data\s+rate|throughput|bandwidth|bitrate)\s+(?:of\s+)?(?:up\s+to\s+|at\s+least\s+|minimum\s+)?\d+(?:\.\d+)?\s*(?:bps|bits?\s+per\s+second|kbps|Mbps|Gbps|bytes?\s+per\s+second|Bps|KBps|MBps|GBps)\b',
                
                # Latency/delay
                r'(?:latency|delay|response\s+time)\s+(?:of\s+)?(?:less\s+than\s+|no\s+more\s+than|maximum\s+)?\d+(?:\.\d+)?\s*(?:ms|millisecond(?:s)?|μs|microsecond(?:s)?|sec|second(?:s)?)\b'
                
                # Storage/memory/capacity
                r'(?:capacity|storage|memory|size)\s+(?:of\s+)?(?:at\s+least\s+|minimum\s+|maximum\s+)?\d+(?:\.\d+)?\s*(?:B|bytes?|KB|MB|GB|TB|PB|kilobytes?|megabytes?|gigabytes?|terabytes?|petabytes?)\b',
                
                # dB/km (signal attenuation)
                r'\d+(?:\.\d+)?\s*dB/km',
                
                # Wavelength (with μm or corrupted ?m)
                r'\d+(?:\.\d+)?\s*(?:μm|µm|\?m|um|micrometers?|microns?)\s+wavelength',
                r'at\s+\d+(?:\.\d+)?\s*(?:μm|µm|\?m|um|micrometers?)',
                
                # ========== RELIABILITY/AVAILABILITY PATTERNS ==========
                r'(?:MTBF|mean\s+time\s+between\s+failures?)\s+(?:of\s+)?(?:at\s+least\s+|minimum\s+)?\d+(?:\.\d+)?\s*(?:hours?|days?|years?|FIT)\b',
                r'(?:MTTR|mean\s+time\s+to\s+repair|mean\s+time\s+to\s+recovery)\s+(?:of\s+)?(?:less\s+than\s+|maximum\s+)?\d+(?:\.\d+)?\s*(?:minutes?|hours?)\b',
                r'(?:availability|uptime|reliability)\s+(?:of\s+)?(?:at\s+least\s+|minimum\s+)?\d+(?:\.\d+)?%\b',
                r'(?:failure\s+rate|probability\s+of\s+failure)\s+(?:of\s+)?(?:less\s+than\s+)?\d+(?:\.\d+)?(?:E[-+]?\d+)?\s*(?:per\s+hour|per\s+year|%|percent)?\b',
                
                # Readiness levels (can be both PERFORMANCE and DESIGN_CONSTRAINTS)
                r'\b(?:TRL|MRL|IRL|SRL)\s+\d+\b',

                # ========== GENERIC COMPARATIVE PATTERNS (Less Specific - Last) ==========
                r'(?:at\s+least|minimum\s+of|no\s+less\s+than|greater\s+than|more\s+than|exceeds?|above)\s+\d+(?:\.\d+)?\s*(?:km|meters?|m|kg|g|hours?|days?|years?)\b',
                r'(?:at\s+most|maximum\s+of|no\s+more\s+than|less\s+than|fewer\s+than|below|under)\s+\d+(?:\.\d+)?\s*(?:km|meters?|m|kg|g|hours?|days?|years?)\b',
                r'(?:equal\s+to|exactly|precisely)\s+\d+(?:\.\d+)?\s*\w+\b',
                r'(?:between|from|ranging\s+from)\s+\d+(?:\.\d+)?\s+(?:and|to)\s+\d+(?:\.\d+)?\s*\w+\b',
                
                # ========== QUALITATIVE PERFORMANCE (Last Resort) ==========
                r'in\s+accordance\s+with\s+(?:specification|standard|requirement|document)\s+[\w\d\-\.]+\b',
                r'per\s+(?:specification|standard|requirement|document)\s+[\w\d\-\.]+\b'
            ],
            
            'TIMING': [
                # ABSOLUTE TIMING
                r'within\s+\d+(?:\.\d+)?\s*(?:ms|milliseconds?|μs|microseconds?|ns|nanoseconds?|sec|seconds?|min|minutes?|hr|hours?|days?)\b',
                r'(?:after|following)\s+\d+(?:\.\d+)?\s*(?:ms|milliseconds?|μs|sec|seconds?|min|minutes?|hr|hours?|days?)\b',
                r'(?:before|prior\s+to)\s+\d+(?:\.\d+)?\s*(?:ms|milliseconds?|sec|seconds?|min|minutes?|hr|hours?|days?)\b',
                r'(?:in|after)\s+(?:less\s+than|no\s+more\s+than)\s+\d+(?:\.\d+)?\s*(?:ms|sec|seconds?|min|minutes?)\b',
                r'(?:no\s+later\s+than|by)\s+\d+(?:\.\d+)?\s*(?:seconds?|minutes?|hours?|days?)\b',
                
                # MISSION/EVENT TIMING
                r'(?:at\s+)?(?:T\s*[+-]\s*\d+|launch\s*[+-]\s*\d+|ignition\s*[+-]\s*\d+)\s*(?:seconds?|minutes?|hours?|days?)\b',
                
                # EVENT-BASED TIMING
                r'(?:upon|on|when)\s+(?:receipt|reception|detection|occurrence|completion|initiation|activation|termination)\s+of\b',
                r'(?:immediately\s+)?(?:upon|after|following|when)\s+(?:system\s+)?(?:startup|boot|initialization|power\s+on|power\s+up|reset|restart)\b',
                r'(?:immediately\s+)?(?:upon|when|after)\s+(?:detection|occurrence)\s+of\s+(?:a\s+)?(?:fault|failure|error|anomaly|condition)\b',
                r'(?:during|throughout|while)\s+(?:flight|mission|operations?|nominal\s+operations?|emergency|fault\s+conditions?|normal\s+mode)\b',
                r'(?:at|upon|during)\s+(?:launch|liftoff|ignition|separation|deployment|orbit\s+insertion|docking|undocking|landing|touchdown)\b',
                r'(?:before|prior\s+to)\s+(?:launch|liftoff|separation|deployment|maneuver|landing|shutdown)\b',
                
                # PERIODIC TIMING
                r'(?:every|each)\s+\d+(?:\.\d+)?\s*(?:ms|milliseconds?|sec|seconds?|min|minutes?|hr|hours?|days?|orbits?|cycles?|revolutions?)\b',
                r'(?:at\s+a\s+rate\s+of|frequency\s+of|at\s+intervals\s+of)\s+\d+(?:\.\d+)?\s*(?:Hz|hertz|times?\s+per\s+second|per\s+minute|per\s+hour|per\s+day)\b',
                r'(?:once\s+per|per)\s+(?:orbit|revolution|cycle|pass|day|hour|minute|second)\b',
                r'(?:continuously|constantly|perpetually)\s+(?:during|throughout|while)\b',
                
                # CONDITIONAL TIMING
                r'(?:while|as\s+long\s+as|during\s+the\s+time\s+that|for\s+the\s+duration\s+of)\s+\w+',
                r'(?:until|up\s+to\s+the\s+point\s+when)\s+\w+',
                
                # REAL-TIME
                r'\b(?:real[- ]?time|instantaneous|immediate|without\s+delay)\b'
            ],
            
            'CONDITION': [
                # OPERATIONAL CONDITIONS
                r'(?:while|during|when)\s+(?:in\s+)?(?:normal|nominal|standby|emergency|fault|degraded|safe|backup|redundant)\s+(?:mode|operation|conditions?|state)\b',
                r'(?:while|during|when)\s+(?:powered|unpowered|on|off|active|inactive|enabled|disabled|energized|de-energized)\b',
                r'(?:while|during|when)\s+(?:the\s+)?(?:system|subsystem|component|unit|equipment)\s+is\s+(?:operating|running|functional|available|operational|online|offline)\b',
                
                # FAULT/FAILURE CONDITIONS
                r'(?:if|when|in\s+the\s+event\s+that|upon)\s+(?:a\s+)?(?:fault|failure|error|anomaly|malfunction|problem)\s+(?:occurs?|is\s+detected|happens)\b',
                r'(?:in\s+the\s+absence\s+of|without|lacking)\s+(?:a\s+)?(?:fault|failure|error|signal|input|power|communication)\b',
                
                # GENERAL CONDITIONS
                r'(?:under|in)\s+(?:all|any|specified|given|defined|normal|abnormal)\s+(?:conditions?|circumstances|scenarios|situations)\b',
                r'(?:subject\s+to|provided\s+that|assuming\s+that|given\s+that)\s+\w+',
                
                # ENVIRONMENTAL CONDITIONS
                r'(?:while|during|when)\s+(?:exposed\s+to|experiencing|subjected\s+to|in)\s+(?:temperature|thermal|vibration|shock|radiation|vacuum|pressure)\s+(?:conditions?|environment)\b',
                r'(?:while|during|when)\s+(?:in\s+)?(?:space|orbit|atmospheric|terrestrial|ground|air|vacuum)\s+(?:environment|conditions?|phase)\b',
                r'(?:while|during|when)\s+(?:in\s+)?(?:sunlight|eclipse|shadow|darkness|daylight|night)\b',
                r'(?:at|under)\s+(?:atmospheric|vacuum|pressurized|unpressurized|ambient)\s+(?:pressure|conditions?)\b',
                
                # MISSION PHASE CONDITIONS
                r'(?:during|throughout|while\s+in)\s+(?:pre-launch|launch|ascent|boost|on-orbit|orbital|re-entry|descent|landing|post-landing|ground)\s+(?:operations?|phase|period|stage)\b',
                r'(?:while|during|when)\s+(?:docked|berthed|attached|separated|free-flying|autonomous|tethered)\b',
                r'(?:during|throughout)\s+(?:mission|flight|operations?|deployment|commissioning|decommissioning|maintenance|servicing)\b',
                
                # SYSTEM STATE CONDITIONS
                r'(?:while|when)\s+(?:receiving|transmitting|processing|computing|storing|monitoring|tracking|controlling)\s+\w+',
                r'(?:while|when)\s+(?:connected\s+to|interfacing\s+with|communicating\s+with|linked\s+to)\s+\w+',
                r'(?:if|when)\s+(?:commanded|requested|triggered|initiated|activated|executed)\s+(?:by|from|via)\s+\w+',
                
                # POWER CONDITIONS
                r'(?:while|when)\s+(?:on\s+)?(?:battery|solar|primary|backup|emergency|redundant)\s+power\b',
                r'(?:during|while)\s+(?:power\s+)?(?:up|down|cycling|switching|transfer|transition)\b',
                r'(?:with|without)\s+(?:external|internal|primary|backup)\s+power\b',
                
                # LOGICAL CONDITIONS
                r'(?:if\s+and\s+only\s+if|unless|except\s+when|provided\s+that)\s+\w+'
            ],
            
            'ENVIRONMENT': [
                # THERMAL ENVIRONMENT (external physical conditions)
                r'temperature\s+(?:range\s+)?(?:of\s+|from\s+|between\s+)?(?:-?\d+(?:\.\d+)?)\s*(?:to\s+(?:-?\d+(?:\.\d+)?)\s*)?(?:°C|°F|K|celsius|fahrenheit|kelvin)\b',
                r'(?:thermal\s+)?(?:cycling|variation|gradient|shock|transient)\s+(?:of\s+|from\s+|between\s+)?\d+(?:\.\d+)?\s*(?:°C|°F|K)\b',
                r'(?:hot|cold|cryogenic|extreme\s+temperature|thermal|heat|freezing)\s+(?:environment|conditions?|exposure)\b',
                r'(?:operating|storage|survival)\s+temperature\s+(?:range)?\b',
                
                # PRESSURE ENVIRONMENT
                r'(?:pressure|vacuum)\s+(?:environment|conditions?)\b',
                r'(?:atmospheric|vacuum|space|ambient|absolute|gauge)\s+pressure\b',
                r'(?:pressurized|unpressurized|vacuum|low\s+pressure|high\s+pressure)\s+(?:environment|conditions?|chamber)\b',
                
                # RADIATION ENVIRONMENT
                r'(?:radiation|particle|cosmic\s+ray|solar\s+particle|Van\s+Allen|ionizing)\s+(?:environment|exposure|flux|dose)\b',
                r'(?:ionizing|non-ionizing|electromagnetic|RF|microwave|gamma|x-ray|alpha|beta|neutron)\s+radiation\b',
                r'(?:total\s+ionizing\s+dose|TID|total\s+dose)\s+(?:environment|exposure)\b',
                r'(?:single\s+event|SEE|SEU|latch-?up)\s+(?:environment|susceptibility|rate)\b',
                
                # SPACE ENVIRONMENT
                r'(?:space|orbital|deep\s+space|interplanetary|cislunar|lunar|martian|planetary)\s+(?:environment|conditions?)\b',
                r'(?:micro)?gravity|weightless(?:ness)?|zero-g|reduced\s+gravity\b',
                r'(?:solar\s+)?(?:wind|plasma|magnetic\s+field|flux)\s+(?:environment|conditions?)\b',
                r'(?:eclipse|sunlight|solar\s+illumination|thermal\s+cycling|day-night\s+cycle)\s+(?:conditions?|environment)\b',
                r'(?:atomic\s+oxygen|AO|plasma|debris|micrometeoroid)\s+(?:environment|exposure)\b',
                
                # HUMIDITY/MOISTURE ENVIRONMENT
                r'(?:humidity|moisture|condensation)\s+(?:environment|conditions?)\b',
                r'(?:wet|dry|humid|arid)\s+(?:environment|conditions?)\b',
                
                # CONTAMINATION ENVIRONMENT
                r'(?:contamination|outgassing|particulate|molecular|chemical)\s+(?:environment|conditions?|requirements?|control)\b',
                r'(?:clean\s+room|cleanroom|sterile|controlled|particle-free)\s+(?:environment|conditions?)\b',
                r'(?:class|ISO)\s+\d+\s+(?:cleanroom|environment)\b',
                
                # CORROSION/CHEMICAL ENVIRONMENT
                r'(?:corrosive|salt\s+spray|salt\s+fog|marine|industrial|chemical)\s+(?:environment|atmosphere|conditions?)\b',
                
                # MECHANICAL ENVIRONMENT (exposure types - NOT performance levels)
                r'(?:vibration|shock|acoustic)\s+(?:environment|conditions?)\b',
                r'(?:random|sinusoidal|pyroshock|seismic|launch|acoustic)\s+(?:environment)\b',
                
                # OPERATIONAL ENVIRONMENT TYPES
                r'(?:ground|pre-launch|launch|ascent|on-orbit|re-entry|landing|surface)\s+(?:environment|conditions?|phase)\b',
                r'(?:transportation|handling|storage|deployment|operational)\s+(?:environment|conditions?)\b',
                
                # EMI/EMC ENVIRONMENT
                r'(?:EMI|EMC|electromagnetic\s+interference|electromagnetic\s+compatibility|ESD|electrostatic\s+discharge)\s+(?:environment|conditions?|requirements?)\b'
            ],
            
            'CONDITION_DURATION': [
                # Duration of CONDITION (for Suitability pattern)
                r'for\s+(?:a\s+)?(?:duration\s+of\s+)?\d+(?:\.\d+)?\s*(?:hours?|days?|years?|cycles?|missions?|orbits?|operations?)\b',
                r'over\s+(?:a\s+period\s+of\s+)?\d+(?:\.\d+)?\s*(?:hours?|days?|years?|missions?)\b',
                r'throughout\s+(?:the\s+)?\d+(?:\.\d+)?\s*(?:hour|day|year|mission|orbit)\b',
                r'during\s+\d+(?:\.\d+)?\s*(?:consecutive\s+)?(?:hours?|days?|cycles?)\b',
            ],
            
            'EXPOSURE_DURATION': [
                # Duration of ENVIRONMENT exposure (for Environment pattern)
                r'for\s+(?:a\s+)?(?:duration\s+of\s+)?\d+(?:\.\d+)?\s*(?:seconds?|minutes?|hours?|days?)\s+(?:of\s+)?exposure\b',
                r'during\s+\d+(?:\.\d+)?\s*(?:seconds?|minutes?|hours?|days?)\s+(?:of\s+)?exposure\b',
                r'after\s+\d+(?:\.\d+)?\s*(?:seconds?|minutes?|hours?|days?)\s+(?:of\s+)?exposure\b',
                r'exposed\s+for\s+\d+(?:\.\d+)?\s*(?:seconds?|minutes?|hours?|days?)\b',
            ],
            
            'CHARACTERISTIC': [
                # Quality attributes (these will be checked via word list in extract function)
                # But add some multi-word patterns here
                r'\b(?:fault\s+tolerance|fault\s+tolerant)\b',
                r'\b(?:mean\s+time\s+between\s+failures?|MTBF)\b',
                r'\b(?:mean\s+time\s+to\s+repair|MTTR)\b',
                r'\b(?:single\s+point\s+of\s+failure|SPOF)\b',
                r'\b(?:fail[-\s]?safe|fail[-\s]?operational)\b',
                r'\b(?:high\s+availability|highly\s+available)\b'
            ],
            
            'DESIGN_CONSTRAINTS': [
                # Negative constraints
                r'\b(?:shall\s+not|must\s+not|cannot|can\s+not|may\s+not)\b',
                r'\b(?:prohibited\s+from|restricted\s+to|limited\s+to|constrained\s+by|confined\s+to)\b',
                
                # Quantitative limits
                r'\b(?:no\s+more\s+than|not\s+to\s+exceed|maximum\s+of)\s+\d+',
                
                # Resource budgets/constraints
                r'\b(?:mass|weight|power|size|volume|energy|cost)\s+(?:limit|constraint|budget|requirement|cap)\b',
                r'\b(?:not\s+to\s+exceed|shall\s+not\s+exceed|limited\s+to)\s+\d+(?:\.\d+)?\s*(?:kg|W|m³|cm³|L)\b',
                
                # Interface constraints
                r'\b(?:interface|mounting|physical|mechanical)\s+(?:constraint|requirement|limitation)\b',
                
                # Material/Manufacturing
                r'\b(?:material|construction|manufacturing|fabrication)\s+(?:restriction|constraint|requirement|limitation)\b',
                
                # Technology Readiness Level (TRL)
                r'\b(?:TRL|Technology\s+Readiness\s+Level)\s+(?:of\s+)?(?:at\s+least\s+|minimum\s+of\s+|≥\s*)?\d+(?:\s+or\s+(?:higher|greater|above))?\b',
                
                # Manufacturing Readiness Level (MRL)
                r'\b(?:MRL|Manufacturing\s+Readiness\s+Level)\s+(?:of\s+)?(?:at\s+least\s+|minimum\s+of\s+|≥\s*)?\d+(?:\s+or\s+(?:higher|greater|above))?\b',
                
                # Standards/Compliance
                r'\b(?:EMI|EMC|electromagnetic\s+(?:interference|compatibility))\s+(?:requirement|constraint|compliance)\b',
                r'\b(?:standard|regulatory|specification)\s+(?:compliance|requirement|constraint)\b',
                r'\bshall\s+comply\s+with\s+(?:MIL-STD|IEEE|ISO|CCSDS|DO-\d+)[-\s]?\d+\b',
                
                # Heritage/COTS
                r'\b(?:heritage\s+hardware|COTS|commercial\s+off[-\s]the[-\s]shelf|existing\s+design)\b'
            ],
            
            'INTERFACE_OUTPUT': [
                r'\b(?:shall\s+)?(?:output|provide|supply|transmit|send|generate|produce)\s+(?:[a-zA-Z\s]+)?(?:signal|data|command|telemetry|status|message)\b',
                r'\b(?:control|command|actuation)\s+(?:signal|output|data)\b',
                r'\b(?:telemetry|status|health|diagnostic)\s+(?:data|information|output)\b',
                r'\b(?:display|indication|alert|alarm|warning|notification)\s+(?:signal|output|message)\b'
            ],
            
            'INTERFACE_INPUT': [
                r'\b(?:shall\s+)?(?:receive|accept|acquire|input|capture|monitor)\s+(?:[a-zA-Z\s]+)?(?:signal|data|command|input|message)\b',
                r'\b(?:sensor|measurement|telemetry)\s+(?:data|input|signal)\b',
                r'\b(?:command|control|configuration)\s+(?:signal|data|input)\b',
                r'\b(?:from|via)\s+(?:the\s+)?(?:operator|user|ground\s+station|external\s+system|interface)\b'
            ],
            
            'EVENT_TRIGGER': [
                r'\b(?:upon|on|when)\s+(?:receipt|reception|detection|occurrence)\s+of\s+(?:a\s+)?(?:[a-zA-Z\s]+)?(?:signal|command|message|event)\b',
                r'\bupon\s+(?:system\s+)?(?:startup|boot|initialization|power\s+on|power\s+up|reset|restart)\b',
                r'\b(?:when|upon)\s+(?:commanded|requested|triggered|initiated)\s+(?:by|from)\b',
                r'\b(?:upon|when)\s+(?:detection|occurrence)\s+of\s+(?:a\s+)?(?:fault|failure|error|anomaly)\b',
                r'\b(?:upon|at)\s+(?:launch|liftoff|separation|deployment|landing|touchdown)\b',
                r'\b(?:when|if)\s+(?:threshold|limit|boundary)\s+(?:exceeded|reached|crossed)\b',
                r'\b(?:upon|when)\s+(?:timer|timeout|countdown)\s+(?:expiration|completion)\b'
            ]
        }


        # Additional word lists for components that need semantic checking

        CHARACTERISTIC_TERMS = {
            # Quality attributes (-ilities)
            "accuracy", "precision", "resolution", "sensitivity", "selectivity",
            "reliability", "availability", "maintainability", "serviceability", "testability",
            "security", "safety", "integrity", "confidentiality", "authenticity",
            "performance", "efficiency", "effectiveness", "throughput", "responsiveness",
            "latency", "bandwidth", "capacity", "scalability", "flexibility",
            "compatibility", "interoperability", "portability", "modularity",
            "robustness", "resilience", "stability", "repeatability", "reproducibility",
            "linearity", "durability", "hardness", "strength", "stiffness",
            "conductivity", "resistivity"
        }


        # Pattern usage examples aligned with INCOSE guidance

        PATTERN_EXAMPLES = {
            'Functional/Performance': {
                'template': 'The {AGENT} shall {FUNCTION} in accordance with {INTERFACE_OUTPUT} with {PERFORMANCE} [and {TIMING} upon {EVENT_TRIGGER} in accordance with {INTERFACE_INPUT}] while in {CONDITION}',
                'example': 'The navigation system shall calculate position with 10-meter accuracy within 5 seconds upon receipt of GPS signals while in normal operation mode',
                'components': ['AGENT', 'FUNCTION', 'PERFORMANCE', 'TIMING', 'EVENT_TRIGGER', 'INTERFACE_INPUT', 'CONDITION']
            },
            
            'Suitability': {
                'template': 'The {AGENT} shall exhibit {CHARACTERISTIC} with {PERFORMANCE} while {CONDITION} [for {CONDITION_DURATION}]',
                'example': 'The spacecraft shall exhibit 99.9% availability with MTBF of 10000 hours during nominal operations for 5 years',
                'components': ['AGENT', 'CHARACTERISTIC', 'PERFORMANCE', 'CONDITION', 'CONDITION_DURATION']
            },
            
            'Environment': {
                'template': 'The {AGENT} shall exhibit {CHARACTERISTIC} during/after exposure to {ENVIRONMENT} [for {EXPOSURE_DURATION}]',
                'example': 'The electronics shall exhibit full functionality after exposure to radiation environment for 1000 hours',
                'components': ['AGENT', 'CHARACTERISTIC', 'ENVIRONMENT', 'EXPOSURE_DURATION']
            },
            
            'Design': {
                'template': 'The {AGENT} shall exhibit {DESIGN_CONSTRAINTS} [in accordance with {PERFORMANCE} while in {CONDITION}]',
                'example': 'The system shall use COTS components with TRL 6 and mass not exceeding 50kg during all mission phases',
                'components': ['AGENT', 'DESIGN_CONSTRAINTS', 'PERFORMANCE', 'CONDITION']
            }
        }

    def extract_incose_components(self, doc) -> Dict[str, Optional[str]]:
        """Extract INCOSE requirement components with enhanced patterns."""
        components = {comp: None for comp in [
            'AGENT', 'FUNCTION', 'CHARACTERISTIC', 'PERFORMANCE', 'CONDITION',
            'ENVIRONMENT', 'TIMING', 'INTERFACE_OUTPUT', 'INTERFACE_INPUT',
            'EVENT_TRIGGER', 'DESIGN_CONSTRAINTS'
        ]}
        
        text = doc.text
        text_lower = text.lower()
        #     # DIAGNOSTIC CODE - ADD THIS
        # if '±' in text or '%' in text or 'dB' in text or '?m' in text:
        #     print(f"\n=== DIAGNOSTIC ===")
        #     print(f"Text: {text[:100]}")
        #     print(f"Has ±: {'±' in text}")
        #     print(f"Has %: {'%' in text}")  
        #     print(f"Has μ: {'μ' in text}")
        #     print(f"Has ?: {'?' in text}")
        #     print(f"Repr: {repr(text[:100])}")
        #     print("==================\n")
    
        # ========== AGENT EXTRACTION (Enhanced NLP) ==========
        for token in doc:
            if token.dep_ == "nsubj" and not token.text.lower() in ["it", "this", "that"]:
                # Get full noun phrase including modifiers
                agent_parts = []
                
                # Get determiners, adjectives, and compounds
                for child in token.children:
                    if child.dep_ in ["det", "amod", "compound", "nmod"]:
                        agent_parts.append((child.i, child.text))
                
                # Add the main noun
                agent_parts.append((token.i, token.text))
                
                # Sort by position and join
                agent_parts.sort(key=lambda x: x[0])
                components['AGENT'] = " ".join([part[1] for part in agent_parts])
                break
        
        # ========== FUNCTION EXTRACTION (Enhanced NLP) ==========
        for token in doc:
            if token.pos_ == "VERB" and token.dep_ == "ROOT":
                function_parts = [token.lemma_]
                
                # Include particles, direct objects, prepositions, and complements
                for child in token.children:
                    if child.dep_ in ["dobj", "prep", "prt", "xcomp"]:
                        function_parts.append(child.text)
                        # Include prepositional objects
                        if child.dep_ == "prep":
                            for grandchild in child.children:
                                if grandchild.dep_ == "pobj":
                                    function_parts.append(grandchild.text)
                
                components['FUNCTION'] = " ".join(function_parts)
                break
        
        # ========== REGEX PATTERN EXTRACTION ==========
        # Extract all components using patterns from self.component_patterns
        for comp_type, patterns in self.component_patterns.items():
            # Skip AGENT and FUNCTION (handled by NLP above)
            if comp_type in ['AGENT', 'FUNCTION']:
                continue
                
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    components[comp_type] = match.group().strip()
                #     # ADD THIS DEBUG
                #     print(f"MATCHED {comp_type}: matched='{match.group()}'")
                # if match and not components[comp_type]:
                #     components[comp_type] = match.group().strip()
                #     # break
        
        # ========== CHARACTERISTIC EXTRACTION (Enhanced with word list) ==========
        if not components['CHARACTERISTIC']:
            # Expanded quality attributes
            characteristic_terms = {
                # Quality attributes
                "accuracy", "precision", "resolution", "sensitivity", "selectivity",
                "reliability", "availability", "maintainability", "serviceability", "testability",
                "security", "safety", "integrity", "confidentiality", "authenticity",
                
                # Performance characteristics
                "performance", "efficiency", "effectiveness", "throughput", "responsiveness",
                "latency", "bandwidth", "capacity", "scalability", "flexibility",
                
                # Compatibility/Interoperability
                "compatibility", "interoperability", "portability", "modularity",
                
                # Robustness
                "robustness", "resilience", "stability", "repeatability", 
                "reproducibility", "linearity",
                
                # Physical characteristics
                "durability", "hardness", "strength", "stiffness", "conductivity", "resistivity"
            }
            
            # Check for single-word characteristics
            for token in doc:
                if token.text.lower() in characteristic_terms:
                    components['CHARACTERISTIC'] = token.text
                    break
        
        return components
    
    def analyze_incose_compliance(self, text: str) -> INCOSEAnalysis:
        """Analyze requirement against INCOSE patterns."""
        doc = self.nlp(text)
        components = self.extract_incose_components(doc)
        
        # Score each pattern
        pattern_scores = []
        for pattern_name, pattern_def in self.patterns.items():
            score = self.score_pattern_match(components, pattern_def)
            missing_req = [comp for comp in pattern_def['required'] if not components.get(comp)]
            missing_opt = [comp for comp in pattern_def['optional'] if not components.get(comp)]
            
            pattern_scores.append({
                'name': pattern_name,
                'score': score,
                'missing_required': missing_req,
                'missing_optional': missing_opt,
                'definition': pattern_def
            })
        
        # Find best matching pattern
        best_match = max(pattern_scores, key=lambda x: x['score'])
        
        # Generate suggestions
        suggestions = self.generate_pattern_suggestions(components, best_match['definition'])
        
        # Create template recommendation
        template_rec = self.create_template_recommendation(components, best_match['definition'])
        
        return INCOSEAnalysis(
            best_pattern=best_match['name'],
            compliance_score=best_match['score'],
            components_found=components,
            missing_required=best_match['missing_required'],
            missing_optional=best_match['missing_optional'],
            suggestions=suggestions,
            template_recommendation=template_rec
        )
    
    def score_pattern_match(self, components: Dict, pattern_def: Dict) -> float:
        """Score how well components match INCOSE pattern."""
        required_found = sum(1 for comp in pattern_def['required'] if components.get(comp))
        optional_found = sum(1 for comp in pattern_def['optional'] if components.get(comp))
        
        required_score = (required_found / len(pattern_def['required'])) * 60
        optional_bonus = (optional_found / len(pattern_def['optional'])) * 40  if pattern_def['optional'] else 0
        
        return min(100, required_score + optional_bonus)
    
    def generate_pattern_suggestions(self, components: Dict, pattern_def: Dict) -> List[str]:
        """Generate improvement suggestions based on missing components."""
        suggestions = []
        
        component_guidance = {
            'AGENT': "Specify the system, subsystem, or component responsible (e.g., 'The navigation system', 'The user interface')",
            'FUNCTION': "Define the specific action or capability (e.g., 'shall calculate', 'shall display', 'shall process')",
            'PERFORMANCE': "Add measurable criteria (e.g., 'within 2 seconds', '±0.1% tolerance')",
            'CONDITION': "Specify operational state (e.g., 'while in normal operation', 'during startup', 'when receiving input')",
            'TIMING': "Add temporal constraints (e.g., 'within 5 seconds', 'upon system startup', 'every 30 minutes')",
            'CHARACTERISTIC': "Define quality attribute (e.g., 'reliability', 'accuracy', 'availability', 'security')",
            'ENVIRONMENT': "Specify environmental conditions (e.g., 'temperature range -40°C to 85°C', 'humidity 0-95%')"
        }
        for comp in pattern_def['required']:
            if not components.get(comp):
                guidance = component_guidance.get(comp, f"Add {comp}")
                suggestions.append(f"Missing {comp}: {guidance}")
        
        return suggestions
    
    def create_template_recommendation(self, components: Dict, pattern_def: Dict) -> str:
        """Create a filled template recommendation."""
        template = pattern_def['template']
        
        # Fill in found components
        for comp, value in components.items():
            if value:
                template = template.replace(f"{{{comp}}}", value)
        
        # Highlight missing components
        for comp in pattern_def['required']:
            if not components.get(comp):
                template = template.replace(f"{{{comp}}}", f"<MISSING {comp}>")
        
        # Remove optional components
        template = re.sub(r'\[.*?\]', '', template)
        
        return template

class SemanticAnalyzer:
    """Semantic quality analyzer for requirements."""
    
    def __init__(self, nlp):
        self.nlp = nlp
        
        # Ambiguous terms
        self.ambiguous_terms = {
            'vague_quantifiers': ['some', 'many', 'few', 'several', 'various'],
            'vague_qualities': ['appropriate', 'adequate', 'suitable', 'proper'],
            'vague_actions': ['handle', 'manage', 'deal', 'support', 'address']
        }
        
        # Subjective terms
        self.subjective_terms = {
            'emotional': ['amazing', 'terrible', 'excellent', 'awful'],
            'subjective': ['good', 'bad', 'best', 'worst', 'better'],
            'uncertainty': ['maybe', 'perhaps', 'possibly', 'might']
        }
    
    def analyze_semantic_quality(self, text: str) -> SemanticAnalysis:
        """Perform semantic analysis on requirement text."""
        doc = self.nlp(text)
        
        # Extract entities
        entity_completeness = self.extract_entities(doc)
        
        # Find ambiguities
        contextual_ambiguities = self.find_contextual_ambiguities(doc)
        
        # Analyze tone
        tone_issues = self.analyze_tone_and_subjectivity(doc)
        
        # Generate suggestions
        suggestions = self.generate_semantic_suggestions(doc)
        
        return SemanticAnalysis(
            similarity_issues=[],  # Will be filled by similarity analysis
            contextual_ambiguities=contextual_ambiguities,
            entity_completeness=entity_completeness,
            tone_issues=tone_issues,
            improvement_suggestions=suggestions
        )
    
    def extract_entities(self, doc) -> Dict[str, List[str]]:
        """Extract semantic entities from requirement."""
        entities = {
            'actors': [],
            'actions': [],
            'objects': [],
            'conditions': [],
            'standards': []
        }
        
        # Extract actors (subjects)
        for token in doc:
            if token.dep_ == "nsubj":
                entities['actors'].append(token.text)
        
        # Extract actions (main verbs)
        for token in doc:
            if token.pos_ == "VERB" and token.dep_ == "ROOT":
                entities['actions'].append(token.lemma_)
        
        # Extract objects
        for token in doc:
            if token.dep_ in ["dobj", "pobj"]:
                entities['objects'].append(token.text)
        
        # Extract conditions
        for token in doc:
            if token.dep_ == "mark" and token.text.lower() in ["if", "when", "while", "during"]:
                # Get the clause
                clause_tokens = list(token.head.subtree)
                condition = " ".join([t.text for t in clause_tokens])
                entities['conditions'].append(condition)
        
        # Extract standards and compliance references
        for ent in doc.ents:
            if ent.label_ == "ORG" and any(std in ent.text.upper() for std in ["ISO", "IEEE", "ANSI", "FIPS"]):
                entities['standards'].append(ent.text)
        return entities
    
    def find_contextual_ambiguities(self, doc) -> List[str]:
        """Find contextual ambiguities in text."""
        ambiguities = []
        
        for category, terms in self.ambiguous_terms.items():
            for token in doc:
                if token.text.lower() in terms:
                    context = self._get_token_context(token, doc)
                    ambiguities.append(f"{token.text} ({category}): {context}")
        
        return ambiguities
    
    def analyze_tone_and_subjectivity(self, doc) -> List[str]:
        """Detect inappropriate tone and subjective language"""
        issues = []
        
        for category, terms in self.subjective_terms.items():
            found_terms = [token.text for token in doc if token.text.lower() in terms]
            if found_terms:
                if category == "emotional":
                    issues.append(f"Emotional language detected: {found_terms} - use neutral, technical terms")
                elif category == "uncertainty":
                    issues.append(f"Uncertainty markers: {found_terms} - requirements should be definitive")
                elif category == "subjective":
                    issues.append(f"Subjective language: {found_terms} - use objective, measurable terms")
        
        return issues
    
    def generate_semantic_suggestions(self, doc) -> List[str]:
        """Generate improvement suggestions based on semantic analysis."""
        suggestions = []
        
        # Check for missing quantification
        has_numbers = any(token.like_num for token in doc)
        if not has_numbers:
            suggestions.append("Add quantitative criteria for verifiability")
        
        # Check for passive voice
        passive_pattern = r'\b(is|are|was|were|been|being)\s+\w+ed\b'
        if re.search(passive_pattern, doc.text, re.IGNORECASE):
            suggestions.append("Use active voice for clarity")

        # Check for vague action verbs
        vague_verbs = ["handle", "manage", "deal with", "work with", "support"]
        for token in doc:
            if token.lemma_ in vague_verbs:
                suggestions.append(f"Replace vague verb '{token.text}' with specific action (e.g., 'process', 'validate', 'calculate')")
        
        # Check for missing error handling
        if any(verb.lemma_ in ["process", "calculate", "validate"] for verb in doc if verb.pos_ == "VERB"):
            if "error" not in doc.text.lower() and "fail" not in doc.text.lower():
                suggestions.append("Consider adding error handling or failure mode specification")
          
        return suggestions
    
    def _get_token_context(self, token, doc, window=3) -> str:
        """Get context around a token."""
        start = max(0, token.i - window)
        end = min(len(doc), token.i + window + 1)
        context_tokens = doc[start:end]
        return " ".join([t.text for t in context_tokens])
    
    def find_similar_requirements(self, requirements_list: List[str], threshold: float = 0.95) -> List[Dict]:
        """Find potentially duplicate requirements using semantic similarity."""
        if not self.nlp.meta.get('vectors', 0):
            return []
        
        docs = [self.nlp(req) for req in requirements_list if req.strip()]
        similarities = []
        
        for i, doc1 in enumerate(docs):
            for j, doc2 in enumerate(docs[i+1:], i+1):
                try:
                    if doc1.vector_norm > 0 and doc2.vector_norm > 0:
                        similarity = doc1.similarity(doc2)
                        if similarity > threshold:
                            similarities.append({
                                'req1_index': i,
                                'req2_index': j,
                                'similarity': float(similarity),
                                'req1_text': requirements_list[i][:100] + "..." if len(requirements_list[i]) > 100 else requirements_list[i],
                                'req2_text': requirements_list[j][:100] + "..." if len(requirements_list[j]) > 100 else requirements_list[j],
                                'issue': f'Potential duplicate (similarity: {similarity:.2f})'
                            })
                except:
                    continue
        
        return similarities

class EnhancedRequirementAnalyzer:
    """Enhanced requirements quality analyzer with INCOSE patterns and advanced NLP."""
    
    def __init__(self, spacy_model: str = "en_core_web_lg", repo_manager=None):
        """Initialize with spaCy model and analyzers."""
        # Initialize utilities
        self.repo_manager = repo_manager or RepositoryStructureManager()
        self.file_handler = SafeFileHandler()
        self.path_resolver = SmartPathResolver()
        
        # Load spaCy model
        try:
            self.nlp = spacy.load(spacy_model)
            logger.info(f"✅ Loaded spaCy model: {spacy_model}")
        except OSError:
            logger.warning(f"Model {spacy_model} not found, using en_core_web_sm")
            self.nlp = spacy.load("en_core_web_sm")
        
        # Initialize specialized analyzers
        self.incose_analyzer = INCOSEPatternAnalyzer(self.nlp)
        self.semantic_analyzer = SemanticAnalyzer(self.nlp)
        
        # Quality criteria
        self.ambiguous_terms = {
            "clarity": {
                "high": {"appropriate", "sufficient", "adequate", "efficient", "reasonable", "acceptable"},
                "medium": {"good", "bad", "proper", "suitable", "normal", "standard"},
                "low": {"nice", "clean", "simple"}
            },
            "ambiguity": {
                "high": {"as needed", "if necessary", "where applicable", "to the extent possible", "as appropriate"},
                "medium": {"typically", "generally", "usually", "often", "sometimes"},
                "low": {"etc", "and so on", "among others"}
            }
        }
        
        self.modal_verbs = {
            'mandatory': ['shall', 'must', 'will'],
            'optional': ['should', 'may'],
            'forbidden': ['shall not', 'must not', 'will not']
        }
        
        self.passive_indicators = [
            r'\b(is|are|was|were|been|being)\s+\w+ed\b',
            r'\b(is|are|was|were|been|being)\s+\w+en\b'
        ]
    
    def analyze_requirement(self, text: str, req_id: str = None) -> Tuple[List[str], QualityMetrics, INCOSEAnalysis, SemanticAnalysis]:
        """Analyze a single requirement with all analysis methods."""
        if pd.isna(text) or not str(text).strip():
            empty_incose = INCOSEAnalysis("", 0, {}, [], [], [], "")
            empty_semantic = SemanticAnalysis([], [], {}, [], [])
            return ["Empty requirement"], QualityMetrics(0, 0, 0, 0, 0, 0, 0, 1, {"critical": 1}), empty_incose, empty_semantic
        
        text = str(text).strip()
        doc = self.nlp(text)
        issues = []
        severity_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        
        # Original quality analysis
        clarity_issues, clarity_score = self._analyze_clarity(doc, text, issues, severity_counts)
        atomicity_issues, atomicity_score = self._analyze_atomicity(doc, issues, severity_counts)
        verifiability_issues, verifiability_score = self._analyze_verifiability(doc, text, issues, severity_counts)
        completeness_issues, completeness_score = self._analyze_completeness(doc, text, issues, severity_counts)
        consistency_score = self._analyze_consistency(doc, issues, severity_counts)
        

        # INCOSE pattern analysis
        incose_analysis = self.incose_analyzer.analyze_incose_compliance(text)
        
        # Semantic analysis
        semantic_analysis = self.semantic_analyzer.analyze_semantic_quality(text)
        
        # Check for implementation details
        self._check_implementation_details(text, issues, severity_counts)

        # Calculate semantic score
        semantic_score = self._calculate_semantic_score(semantic_analysis)
        
        # Build comprehensive result
        has_critical = severity_counts['critical'] > 0
        base_score = (
            clarity_score * 0.2 + 
            completeness_score * 0.2 + 
            verifiability_score * 0.25 +  # Increase weight for verifiability
            atomicity_score * 0.15 +      # Decrease weight for atomicity
            consistency_score * 0.2
        )
        if has_critical:
        # Critical failures cap the score at 40 (POOR grade max)
            quality_score = min(base_score, 40)
        else:
            quality_score = base_score

        # Create metrics
        metrics = QualityMetrics(
            quality_score= quality_score,
            clarity_score=clarity_score,
            completeness_score=completeness_score,
            verifiability_score=verifiability_score,
            atomicity_score=atomicity_score,
            consistency_score=consistency_score,
            incose_compliance_score=incose_analysis.compliance_score,
            semantic_quality_score=semantic_score,
            total_issues=len(issues),
            severity_breakdown=severity_counts
        )

        return issues, metrics, incose_analysis, semantic_analysis
    
    def _analyze_clarity(self, doc, text: str, issues: List[str], severity_counts: Dict[str, int]) -> Tuple[int, float]:
        """Analyze clarity."""
        clarity_issues = 0
        
        # Ambiguous terms
        for token in doc:
            if token.pos_ in {"ADJ", "ADV"}:
                token_lower = token.text.lower()
                for severity, terms in self.ambiguous_terms["clarity"].items():
                    if token_lower in terms:
                        issues.append(f"Clarity ({severity}): ambiguous term '{token.text}'")
                        severity_counts[severity] += 1
                        clarity_issues += 1
        
        # Passive voice
        passive_issues = []
        for pattern in self.passive_indicators:
            if re.search(pattern, text, re.IGNORECASE):
                issues.append(f"Clarity (medium): Passive voice detected: {', '.join(set(passive_issues))}")
                severity_counts["medium"] += 1
                clarity_issues += 1
                break
        
        # # Sentence complexity
        # sentences = list(doc.sents)
        # for sent in sentences:
        #     if len([t for t in sent if not t.is_punct]) > 25:
        #         issues.append("Clarity (medium): Complex sentence")
        #         severity_counts["medium"] += 1
        #         clarity_issues += 1
        
        # Readability
        readability = self._calculate_readability_score(doc)
        if readability < 30:
            issues.append(f"Clarity (medium): low readability score ({readability:.1f}/100)")
            severity_counts["medium"] += 1
            clarity_issues += 1
        
        clarity_score = max(0, 100 - (clarity_issues * 20))

        return clarity_issues, clarity_score
    
    def _analyze_completeness(self, doc, text: str, issues: List[str], severity_counts: Dict[str, int]) -> Tuple[int, float]:
        """Analyze completeness."""
        completeness_issues = 0
        
        # Check for modal verbs
        text_lower = text.lower()
        has_modal = any(modal in text_lower for modals in self.modal_verbs.values() for modal in modals)
        
        if not has_modal:
            issues.append("Completeness (high): Missing modal verb (shall/must/should)")
            severity_counts["high"] += 1
            completeness_issues += 1
        
        # Check for subject
        has_subject = any(token.dep_ == "nsubj" for token in doc)
        if not has_subject:
            issues.append("Completeness (critical): Missing subject/actor")
            severity_counts["critical"] += 1
            completeness_issues += 1
        
        # Check for action
        has_verb = any(token.pos_ == "VERB" for token in doc)
        if not has_verb:
            issues.append("Completeness (critical): Missing action/verb")
            severity_counts["critical"] += 1
            completeness_issues += 1
        
        # Length check
        word_count = len([token for token in doc if token.is_alpha])
        if word_count < 5:
            issues.append("Completeness (medium): Too brief")
            severity_counts["medium"] += 1
            completeness_issues += 1
        
        completeness_score = max(0, 100 - (completeness_issues * 40))
        return completeness_issues, completeness_score
    
    def _analyze_verifiability(self, doc, text: str, issues: List[str], severity_counts: Dict[str, int]) -> Tuple[int, float]:
        """Analyze verifiability."""
        verifiability_issues = 0
        
        # Check for measurable entities
        measurable_entities = [ent for ent in doc.ents if ent.label_ in {"CARDINAL", "QUANTITY", "PERCENT", "TIME", "MONEY"}]
        
        # Enhanced patterns for measurable criteria
        verifiability_patterns = [
            r'\b\d+(?:\.\d+)?\s*(?:mm|cm|m|km|in|ft|°C|°F|K|Hz|rpm|V|A|W|Pa|psi|g|kg|lb)\b',
            r'\b\d+(?:\.\d+)?%\b',
            r'\b\d+(?:\.\d+)?[eE][+-]?\d+\s*[a-zA-Z]+\b',
            r'\b(?:less|more|greater|equal)\s+than\s+\d+(?:\.\d+)?',
            r'\b(?:within|±|plus|minus)\s*\d+(?:\.\d+)?',
            r'\b(?:at\s+least|at\s+most|exactly|up\s+to)\s+\d+',
            r'\b(?:ISO|IEEE|ANSI|ASTM|MIL-STD|DO-\d+|IEC|FIPS)\s*[-]?\s*\d+',
            r'\b(?:accuracy|precision|tolerance|error)\s+(?:of\s+)?[±]?\d+(?:\.\d+)?'
        ]
        
        has_verifiable_criteria = bool(measurable_entities)
        
        if not has_verifiable_criteria:
            text_lower = text.lower()
            for pattern in verifiability_patterns:
                if re.search(pattern, text_lower, re.IGNORECASE):
                    has_verifiable_criteria = True
                    break
        
        if not has_verifiable_criteria:
            issues.append("Verifiability (critical): no measurable criteria found")
            severity_counts["critical"] += 1
            verifiability_issues += 1
        
        verifiability_score = max(0, 100 - (verifiability_issues * 40))
        return verifiability_issues, verifiability_score
    
    def _analyze_atomicity(self, doc, issues: List[str], severity_counts: Dict[str, int]) -> Tuple[int, float]:
        """Analyze atomicity"""
        atomicity_issues = 0
        conjunction_count = sum(1 for token in doc if token.dep_ == "conj")
        
        if conjunction_count > 2:
            issues.append(f"Atomicity (high): multiple conjunctions ({conjunction_count}) suggest compound requirements")
            severity_counts["high"] += 1
            atomicity_issues += 1
        elif conjunction_count > 0:
            issues.append(f"Atomicity (medium): contains {conjunction_count} conjunction(s)")
            severity_counts["medium"] += 1
            atomicity_issues += 1
        
        # Length analysis
        word_count = len([token for token in doc if token.is_alpha])
        if word_count > 50:
            issues.append("Atomicity (low): requirement may be too long (consider splitting)")
            severity_counts["low"] += 1
        
        atomicity_score = max(0, 100 - (atomicity_issues * 30))
        return atomicity_issues, atomicity_score
        
    def _analyze_consistency(self, doc, issues: List[str], severity_counts: Dict[str, int]) -> float:
        """Analyze consistency."""
        # For now, just check if modal verb is present
        modal_found = False
        for strength, verbs in self.modal_verbs.items():
            for token in doc:
                if token.text.lower() in verbs:
                    modal_found = True
                    break
            if not modal_found:
                issues.append("Consistency (high): No modal verb for requirement strength")
                severity_counts["high"] += 1
                return 50  
        return 100
    
    def _check_implementation_details(self, text: str, issues: List[str], severity_counts: Dict[str, int]):
        """Check for implementation details"""
        implementation_indicators = [
            r'\b(?:implement|deploy|install|configure|setup|utilize|employ)\b',
            r'\busing\s+(?:a|an|the)?\s*[A-Z][a-zA-Z]*(?:\s+[A-Z][a-zA-Z]*)*',
            r'\b(?:via|through)\s+(?:a|an|the)?\s*[a-zA-Z]+',
            r'\bby\s+means\s+of\b',
            r'\bwith\s+(?:a|an|the)?\s*[A-Z][a-zA-Z]*(?:\s+[A-Z][a-zA-Z]*)*(?:\s+(?:database|server|protocol))?',
        ]
        
        # Exclude verification language
        verification_patterns = [
            r'\bas\s+(?:verified|validated|demonstrated|tested)\s+(?:through|by|via)',
            r'\bsubject\s+to\s+(?:verification|validation|testing)',
        ]
        
        text_lower = text.lower()
        has_implementation = False
        
        for pattern in implementation_indicators:
            if re.search(pattern, text_lower):
                is_verification = any(re.search(ver_pattern, text_lower) for ver_pattern in verification_patterns)
                if not is_verification:
                    has_implementation = True
                    break
        
        if has_implementation:
            issues.append("Design (high): Contains implementation details")
            severity_counts["high"] += 1

    def _calculate_readability_score(self, doc) -> float:
        """Calculate readability score"""
        sentences = list(doc.sents)
        if not sentences:
            return 0.0
        
        avg_sentence_length = len([token for token in doc if not token.is_punct]) / len(sentences)
        complex_words = sum(1 for token in doc if len(token.text) > 6 and token.is_alpha)
        total_words = sum(1 for token in doc if token.is_alpha)
        complex_ratio = complex_words / max(total_words, 1)
        
        score = max(0, 100 - (avg_sentence_length * 2 + complex_ratio * 50))
        return min(score, 100)
   
    def _calculate_semantic_score(self, semantic_analysis: SemanticAnalysis) -> float:
        """Calculate semantic quality score."""
        base_score = 100
        
        # Penalize issues
        penalty = len(semantic_analysis.contextual_ambiguities) * 10
        penalty += len(semantic_analysis.tone_issues) * 5
        
        # Bonus for good entity completeness
        entity_bonus = 0
        if semantic_analysis.entity_completeness.get('actors'):
            entity_bonus += 5
        if semantic_analysis.entity_completeness.get('actions'):
            entity_bonus += 5
        if semantic_analysis.entity_completeness.get('objects'):
            entity_bonus += 5
        
        return max(0, min(100, base_score - penalty + entity_bonus))
    
    def analyze_file(self, input_file: str = "requirements.csv", 
                    output_file: str = None,
                    requirement_column: str = "Requirement Text",
                    excel_report: bool = True) -> pd.DataFrame:
        """Analyze requirements file with enhanced output."""
        logger.info(f"Starting enhanced analysis of {input_file}")
        
        # Resolve file path
        resolved_paths = self.path_resolver.resolve_input_files({'requirements': input_file})
        input_file_path = resolved_paths['requirements']
        
        if not Path(input_file_path).exists():
            raise FileNotFoundError(f"Could not find requirements file: {input_file}")
        
        # Read file
        df = self.file_handler.safe_read_csv(input_file_path)
        
        if requirement_column not in df.columns:
            available_cols = list(df.columns)
            logger.error(f"Column '{requirement_column}' not found. Available: {available_cols}")
            raise ValueError(f"Column '{requirement_column}' not found in CSV")
        
        # Ensure we have required columns per conventions.md
        if 'ID' not in df.columns:
            df['ID'] = [f"REQ_{i:04d}" for i in range(len(df))]
        
        # Get Requirement Name if available
        req_name_col = 'Requirement Name' if 'Requirement Name' in df.columns else None
        
        df = df.fillna({requirement_column: ""})
        logger.info(f"Analyzing {len(df)} requirements...")
        
        # Analyze each requirement
        all_analysis_results = []
        
        for idx, row in df.iterrows():
            req_id = row['ID']
            req_text = row[requirement_column]
            req_name = row[req_name_col] if req_name_col else ""
            
            issues, metrics, incose_analysis, semantic_analysis = self.analyze_requirement(req_text, req_id)
            
            # # Build comprehensive result
            # has_critical = metrics.severity_breakdown['critical'] > 0
            # base_score = (
            #     metrics.clarity_score * 0.2 + 
            #     metrics.completeness_score * 0.2 + 
            #     metrics.verifiability_score * 0.25 +  # Increase weight for verifiability
            #     metrics.atomicity_score * 0.15 +      # Decrease weight for atomicity
            #     metrics.consistency_score * 0.2
            # )
            # if has_critical:
            # # Critical failures cap the score at 40 (POOR grade max)
            #     quality_score = min(base_score, 40)
            # else:
            #     quality_score = base_score

            result = {

                # Core columns
                'ID': req_id,
                'Requirement Name': req_name,
                'Requirement Text': req_text,
                
                # Quality scores
                'Quality_Score': metrics.quality_score,
                'Quality_Grade': self._get_grade(metrics.quality_score),

                
                # Quality breakdown
                'Clarity_Score': metrics.clarity_score,
                'Completeness_Score': metrics.completeness_score,
                'Verifiability_Score': metrics.verifiability_score,
                'Atomicity_Score': metrics.atomicity_score,
                'Consistency_Score': metrics.consistency_score,
                
                # Issues
                'Total_Issues': metrics.total_issues,
                'Critical_Issues': metrics.severity_breakdown['critical'],
                'High_Issues': metrics.severity_breakdown['high'],
                'Medium_Issues': metrics.severity_breakdown['medium'],
                'Low_Issues': metrics.severity_breakdown['low'],
                'Issue_Details': '; '.join(issues),
                
                # INCOSE analysis
                'INCOSE_Compliance_Score': incose_analysis.compliance_score,
                'INCOSE_Best_Pattern': incose_analysis.best_pattern,
                'INCOSE_Missing_Required': ', '.join(incose_analysis.missing_required) if incose_analysis.missing_required else 'None',
                'INCOSE_Missing_Optional': ', '.join(incose_analysis.missing_optional) if incose_analysis.missing_optional else 'None',
                'INCOSE_Template': incose_analysis.template_recommendation,
                'INCOSE_Suggestions': '; '.join(incose_analysis.suggestions),
                
                # INCOSE Components found
                'Has_Agent': 'Yes' if incose_analysis.components_found.get('AGENT') else 'No',
                'Has_Function': 'Yes' if incose_analysis.components_found.get('FUNCTION') else 'No',
                'Has_Performance': 'Yes' if incose_analysis.components_found.get('PERFORMANCE') else 'No',
                'Has_Condition': 'Yes' if incose_analysis.components_found.get('CONDITION') else 'No',
                
                # Semantic analysis
                'Semantic_Quality_Score': metrics.semantic_quality_score,
                'Actors_Found': ', '.join(semantic_analysis.entity_completeness.get('actors', [])) if semantic_analysis.entity_completeness.get('actors') else 'None',
                'Actions_Found': ', '.join(semantic_analysis.entity_completeness.get('actions', [])) if semantic_analysis.entity_completeness.get('actions') else 'None',
                'Objects_Found': ', '.join(semantic_analysis.entity_completeness.get('objects', [])) if semantic_analysis.entity_completeness.get('objects') else 'None',
                'Conditions_Found': ', '.join(semantic_analysis.entity_completeness.get('conditions', [])) if semantic_analysis.entity_completeness.get('conditions') else 'None',
                
                'Ambiguous_Terms': '; '.join(semantic_analysis.contextual_ambiguities) if semantic_analysis.contextual_ambiguities else 'None',
                'Tone_Issues': '; '.join(semantic_analysis.tone_issues) if semantic_analysis.tone_issues else 'None',
                'Semantic_Suggestions': '; '.join(semantic_analysis.improvement_suggestions) if semantic_analysis.improvement_suggestions else 'None',
                
                # Will be filled after similarity analysis
                'Similar_Requirements': 0,
                'Most_Similar_ID': 'None',
                'Max_Similarity': 0.0,
                'Duplicate_Group': 'UNIQUE'
            }
            
            all_analysis_results.append(result)
            
            if (idx + 1) % 50 == 0:
                logger.info(f"Processed {idx + 1}/{len(df)} requirements")
        
        # Create results DataFrame
        results_df = pd.DataFrame(all_analysis_results)
        
        # Find similar requirements
        logger.info("Finding similar requirements...")
        req_texts = df[requirement_column].tolist()
        similarity_results = self.semantic_analyzer.find_similar_requirements(req_texts)
        
        # Update similarity information
        similarity_map = {}
        for sim_result in similarity_results:
            idx1, idx2 = sim_result['req1_index'], sim_result['req2_index']
            
            # Map to IDs
            id1 = results_df.iloc[idx1]['ID']
            id2 = results_df.iloc[idx2]['ID']
            
            if id1 not in similarity_map:
                similarity_map[id1] = []
            if id2 not in similarity_map:
                similarity_map[id2] = []
            
            similarity_map[id1].append((id2, sim_result['similarity']))
            similarity_map[id2].append((id1, sim_result['similarity']))
        
        # Update similarity columns
        for idx, row in results_df.iterrows():
            req_id = row['ID']
            if req_id in similarity_map:
                similar_reqs = similarity_map[req_id]
                results_df.at[idx, 'Similar_Requirements'] = len(similar_reqs)
                if similar_reqs:
                    most_similar = max(similar_reqs, key=lambda x: x[1])
                    results_df.at[idx, 'Most_Similar_ID'] = most_similar[0]
                    results_df.at[idx, 'Max_Similarity'] = most_similar[1]
        
        # Add duplicate groups
        results_df = self._add_duplicate_groups(results_df, similarity_map)
        
        # # Save CSV results
        # if output_file is None:
        #     output_file = Path(input_file_path).stem + "_quality_analysis"
        
        # csv_path = Path(output_file).with_suffix('.csv')
        # results_df.to_csv(csv_path, index=False)
        # logger.info(f"Saved CSV results to {csv_path}")
        
        self._save_enhanced_results(results_df, input_file_path, output_file, excel_report)
        # Create Excel report if requested
        if excel_report:
            excel_path = self._create_excel_report(results_df, output_file)
            logger.info(f"Created Excel report: {excel_path}")
        
        # Print summary
        self._print_summary(results_df)
        
        return results_df
    
    def _save_enhanced_results(self, df: pd.DataFrame, input_file_path: str, 
                              output_file: str, excel_report: bool) -> str:
        """Save enhanced analysis results"""
        
        # Determine output file path
        if not output_file:
            input_stem = Path(input_file_path).stem
            output_file = self.file_handler.get_structured_path(
                'quality_analysis', 
                f"{input_stem}_quality_report.csv"
            )
        
        # Save CSV
        try:
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
            df.to_json(Path(output_file).with_suffix('.json'), orient='records', indent=2, force_ascii=False)
            logger.info(f"Enhanced CSV analysis saved to '{output_file}'")
        except Exception as e:
            logger.error(f"Failed to save CSV: {e}")
            raise
        
        # # Save enhanced summary
        # summary_report = self._generate_enhanced_summary(df)
        # summary_file = str(Path(output_file).with_suffix("")) + "_enhanced_summary.json"
        
        # try:
        #     with open(summary_file, "w", encoding='utf-8') as f:
        #         json.dump(summary_report, f, indent=2, default=str, ensure_ascii=False)
        #     logger.info(f"Enhanced summary saved to '{summary_file}'")
        # except Exception as e:
        #     logger.warning(f"Could not save enhanced summary: {e}")
        
        # # Create enhanced Excel report
        # if excel_report:
        #     try:
        #         excel_path = self._create_enhanced_excel_report(df)
        #         logger.info(f"Enhanced Excel report created: {excel_path}")
        #     except Exception as e:
        #         logger.error(f"Failed to create enhanced Excel report: {e}")
        
        return output_file
    
    def _get_grade(self, score: float) -> str:
        """Convert score to grade."""
        if score >= 90:
            return 'EXCELLENT'
        elif score >= 75:
            return 'GOOD'
        elif score >= 60:
            return 'FAIR'
        elif score >= 40:
            return 'POOR'
        else:
            return 'CRITICAL'
    
    def _add_duplicate_groups(self, df: pd.DataFrame, similarity_map: Dict[str, List[Tuple[str, float]]]) -> pd.DataFrame:
        """Add duplicate group IDs."""
        groups = {}
        group_id = 1
        processed = set()
        
        for req_id, similar_reqs in similarity_map.items():
            if req_id in processed:
                continue
            
            # Find all requirements with >95% similarity
            group_members = {req_id}
            for sim_id, sim_score in similar_reqs:
                if sim_score >= 0.95:
                    group_members.add(sim_id)
            
            if len(group_members) > 1:
                for member in group_members:
                    groups[member] = f"DUP_GROUP_{group_id:03d}"
                    processed.add(member)
                group_id += 1
        
        # Map to DataFrame
        df['Duplicate_Group'] = df['ID'].map(groups).fillna('UNIQUE')
        
        return df
    
    def _create_excel_report(self, df: pd.DataFrame, base_filename: str) -> Path:
        """Create comprehensive 4-tab Excel report with enhanced formatting."""
        output_file = self.file_handler.get_structured_path(
            'quality_analysis', 
            "requirements_quality_report.xlsx"
        )
        
        excel_path = Path(output_file)
        excel_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Tab 1: Summary
            summary_df = self._create_summary_tab(df)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Tab 2: Quality Grading
            quality_cols = [
                'ID', 'Requirement Name', 'Requirement Text',
                'Quality_Score', 'Quality_Grade',
                'Clarity_Score', 'Completeness_Score', 'Verifiability_Score', 
                'Atomicity_Score', 'Consistency_Score',
                'Total_Issues', 'Critical_Issues', 'High_Issues', 'Medium_Issues', 'Low_Issues',
                'Issue_Details'
            ]
            quality_df = df[quality_cols]
            quality_df.to_excel(writer, sheet_name='Quality Grading', index=False)
            
            # Tab 3: INCOSE Analysis
            incose_cols = [
                'ID', 'Requirement Name', 'Requirement Text',
                'INCOSE_Compliance_Score', 'INCOSE_Best_Pattern',
                'Has_Agent', 'Has_Function', 'Has_Performance', 'Has_Condition',
                'INCOSE_Missing_Required', 'INCOSE_Missing_Optional', 'INCOSE_Suggestions'
            ]
            incose_df = df[incose_cols]
            incose_df.to_excel(writer, sheet_name='INCOSE Analysis', index=False)
            
            # Tab 4: Semantic Analysis
            semantic_cols = [
                'ID', 'Requirement Name', 'Requirement Text',
                'Semantic_Quality_Score',
                'Actors_Found', 'Actions_Found', 'Objects_Found', 'Conditions_Found',
                'Ambiguous_Terms', 'Tone_Issues',
                'Similar_Requirements', 'Most_Similar_ID', 'Max_Similarity', 'Duplicate_Group',
                'Semantic_Suggestions'
            ]
            semantic_df = df[semantic_cols]
            semantic_df.to_excel(writer, sheet_name='Semantic Analysis', index=False)
            
            # Apply enhanced formatting
            self._format_excel_workbook(writer)
        
        return excel_path
    
    def _format_excel_workbook(self, writer):
        """Apply enhanced formatting with tables, auto-sizing, and conditional formatting."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
            
            workbook = writer.book
            
            # Define formatting styles
            header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Grade colors for conditional formatting
            grade_fills = {
                'EXCELLENT': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                'GOOD': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                'FAIR': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                'POOR': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                'CRITICAL': PatternFill(start_color="FF5050", end_color="FF5050", fill_type="solid")
            }
            
            # Format each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                try:
                    # Special handling for Summary sheet - NO TABLE, just formatting
                    if sheet_name == 'Summary':
                        # Format headers
                        if worksheet.max_row > 0:
                            for cell in worksheet[1]:
                                if cell.value:
                                    cell.font = Font(bold=True, size=12)
                                    cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                                    cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                        # Auto-fit columns for Summary
                        for column_cells in worksheet.columns:
                            column_letter = get_column_letter(column_cells[0].column)
                            
                            # Calculate max width
                            max_length = 0
                            for cell in column_cells:
                                try:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                except:
                                    pass
                            
                            # Set width with reasonable limits
                            adjusted_width = min(max_length + 2, 150)
                            worksheet.column_dimensions[column_letter].width = max(adjusted_width, 15)
                        
                        # Add styling to section headers
                        for row_idx in range(2, worksheet.max_row + 1):
                            section_cell = worksheet.cell(row_idx, 1)
                            if section_cell.value and section_cell.value.strip() and worksheet.cell(row_idx, 2).value == '':
                                # This is a section header
                                for col_idx in range(1, 5):
                                    cell = worksheet.cell(row_idx, col_idx)
                                    cell.font = Font(bold=True, size=11)
                                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        
                        logger.info("✓ Formatted Summary (no table)")
                    
                    # Create Excel tables for data tabs
                    else:
                        # Determine the table range
                        max_row = worksheet.max_row
                        max_col = worksheet.max_column
                        
                        if max_row > 1 and max_col > 0:  # Ensure there's data beyond headers
                            # Create table reference
                            table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
                            
                            # Clean table name
                            table_name = sheet_name.replace(' ', '_')
                            
                            # Create and add table
                            tab = Table(displayName=table_name, ref=table_ref)
                            style = TableStyleInfo(
                                name="TableStyleMedium9", 
                                showFirstColumn=False,
                                showLastColumn=False, 
                                showRowStripes=True, 
                                showColumnStripes=False
                            )
                            tab.tableStyleInfo = style
                            worksheet.add_table(tab)
                            logger.info(f"✓ Created table for {sheet_name}")
                        
                        # Apply column-specific formatting
                        for column_cells in worksheet.columns:
                            column_letter = get_column_letter(column_cells[0].column)
                            
                            # Skip empty columns
                            if not any(cell.value for cell in column_cells):
                                continue
                            
                            # Get column header name
                            header_value = column_cells[0].value
                            column_name = str(header_value) if header_value else ""
                            
                            # Calculate max width based on content
                            max_length = 0
                            for cell in column_cells:
                                try:
                                    if cell.value:
                                        cell_value = str(cell.value)
                                        lines = cell_value.split('\n')
                                        max_line_length = max(len(line) for line in lines) if lines else len(cell_value)
                                        max_length = max(max_length, max_line_length)
                                except:
                                    pass
                            
                            # Adjust width with a multiplier
                            adjusted_width = min(max_length * 1.1 + 2, 100)
                            
                            # Apply column-specific constraints
                            if 'Requirement Text' in column_name or 'Requirement_Text' in column_name:
                                # Requirement Text: Wide with wrapping
                                worksheet.column_dimensions[column_letter].width = max(50, min(adjusted_width, 80))
                                
                                # Apply wrap text
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:
                                        cell.alignment = Alignment(
                                            wrap_text=True, 
                                            vertical='top', 
                                            horizontal='left'
                                        )
                            
                            elif 'ID' in column_name:
                                # ID columns: Narrow
                                worksheet.column_dimensions[column_letter].width = max(8, min(adjusted_width, 15))
                                
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                            
                            elif 'Score' in column_name:
                                # Score columns: Medium width, centered
                                worksheet.column_dimensions[column_letter].width = max(12, min(adjusted_width, 30))
                                
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                        # Format scores
                                        try:
                                            if isinstance(cell.value, (int, float)):
                                                if 'INCOSE' in column_name:
                                                    cell.number_format = '0.0"%"'
                                                else:
                                                    cell.number_format = '0.0'
                                        except:
                                            pass
                            
                            elif 'Grade' in column_name:
                                # Grade column: Apply conditional formatting
                                worksheet.column_dimensions[column_letter].width = 15
                                
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                        # Apply grade colors
                                        if cell.value in grade_fills:
                                            cell.fill = grade_fills[cell.value]
                                            if cell.value in ['POOR', 'CRITICAL']:
                                                cell.font = Font(color="FFFFFF", bold=True)
                            
                            elif any(keyword in column_name for keyword in ['Details', 'Suggestions', 'Missing', 'Found']):
                                # Text detail columns: Wide
                                worksheet.column_dimensions[column_letter].width = max(30, min(adjusted_width, 60))
                                
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:
                                        cell.alignment = Alignment(
                                            horizontal='left', 
                                            vertical='top', 
                                            wrap_text=True
                                        )
                            
                            elif 'Name' in column_name:
                                # Name columns: Medium width
                                worksheet.column_dimensions[column_letter].width = max(20, min(adjusted_width, 40))
                                
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1 and cell.value:
                                        cell.alignment = Alignment(horizontal='left', vertical='center')
                            
                            else:
                                # Default columns
                                worksheet.column_dimensions[column_letter].width = max(12, min(adjusted_width, 40))
                                
                                for row_idx, cell in enumerate(column_cells, 1):
                                    if row_idx > 1:
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        
                        # Set intelligent row heights
                        for row_idx in range(2, worksheet.max_row + 1):
                            has_wrapping_text = False
                            max_text_length = 0
                            
                            for cell in worksheet[row_idx]:
                                if cell.value:
                                    cell_text = str(cell.value)
                                    header_cell = worksheet.cell(1, cell.column)
                                    column_name = str(header_cell.value) if header_cell.value else ""
                                    
                                    if 'Requirement Text' in column_name and len(cell_text) > 50:
                                        has_wrapping_text = True
                                        max_text_length = max(max_text_length, len(cell_text))
                            
                            # Set row height based on content
                            if has_wrapping_text:
                                estimated_lines = max(2, min(10, max_text_length // 80))
                                worksheet.row_dimensions[row_idx].height = 30 + (estimated_lines * 12)
                            else:
                                worksheet.row_dimensions[row_idx].height = 20
                    
                    # Set header row height (all sheets)
                    if worksheet.max_row >= 1:
                        worksheet.row_dimensions[1].height = 25
                    
                    # Freeze panes (all sheets except Summary)
                    if sheet_name != 'Summary':
                        worksheet.freeze_panes = worksheet['A2']
                    
                    # Add borders to all cells with data
                    thin_border = Border(
                        left=Side(style='thin', color='E0E0E0'),
                        right=Side(style='thin', color='E0E0E0'),
                        top=Side(style='thin', color='E0E0E0'),
                        bottom=Side(style='thin', color='E0E0E0')
                    )
                    
                    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                                min_col=1, max_col=worksheet.max_column):
                        for cell in row:
                            if cell.value is not None:
                                if cell.row == 1:  # Header row
                                    cell.border = Border(
                                        left=Side(style='medium'),
                                        right=Side(style='medium'),
                                        top=Side(style='medium'),
                                        bottom=Side(style='medium')
                                    )
                                else:
                                    cell.border = thin_border
                    
                    logger.info(f"✓ Formatted sheet: {sheet_name}")
                
                except Exception as e:
                    logger.warning(f"Could not format sheet {sheet_name}: {e}")
                    import traceback
                    traceback.print_exc()
        
        except ImportError:
            logger.warning("openpyxl styling not available - basic formatting applied")
        except Exception as e:
            logger.warning(f"Workbook formatting failed: {e}")
            import traceback
            traceback.print_exc()

    def _create_summary_tab(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create comprehensive summary tab with methodology."""
        summary_data = []
        
        # Overall Statistics
        summary_data.extend([
            {'Section': 'OVERALL STATISTICS', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Total Requirements', 'Value': len(df), 'Details': ''},
            {'Section': '', 'Metric': 'Average Quality Score', 'Value': f"{df['Quality_Score'].mean():.1f}/100", 'Details': ''},
            {'Section': '', 'Metric': 'Average INCOSE Compliance', 'Value': f"{df['INCOSE_Compliance_Score'].mean():.1f}%", 'Details': ''},
            {'Section': '', 'Metric': 'Average Semantic Score', 'Value': f"{df['Semantic_Quality_Score'].mean():.1f}/100", 'Details': ''},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # Grade Distribution
        summary_data.append({'Section': 'GRADE DISTRIBUTION', 'Metric': '', 'Value': '', 'Details': ''})
        grades = df['Quality_Grade'].value_counts()
        for grade in ['EXCELLENT', 'GOOD', 'FAIR', 'POOR', 'CRITICAL']:
            count = grades.get(grade, 0)
            pct = count / len(df) * 100 if len(df) > 0 else 0
            summary_data.append({
                'Section': '',
                'Metric': f'{grade} Requirements',
                'Value': f'{count} ({pct:.1f}%)',
                'Details': self._get_grade_description(grade)
            })
        summary_data.append({'Section': '', 'Metric': '', 'Value': '', 'Details': ''})
        
        # Top Issues
        summary_data.append({'Section': 'TOP ISSUES', 'Metric': '', 'Value': '', 'Details': ''})
        all_issues = []
        for issues_text in df['Issue_Details']:
            if issues_text and str(issues_text) != 'nan':
                issues = issues_text.split(';')
                for issue in issues:
                    if ':' in issue:
                        issue_type = issue.split(':')[0].strip()
                        all_issues.append(issue_type)
        
        if all_issues:
            issue_counts = Counter(all_issues)
            for issue_type, count in issue_counts.most_common(5):
                summary_data.append({
                    'Section': '',
                    'Metric': issue_type,
                    'Value': count,
                    'Details': ''
                })
        summary_data.append({'Section': '', 'Metric': '', 'Value': '', 'Details': ''})
        
        # Duplicates
        dup_count = len(df[df['Duplicate_Group'] != 'UNIQUE'])
        summary_data.extend([
            {'Section': 'DUPLICATE ANALYSIS', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Potential Duplicates', 'Value': dup_count, 'Details': 'Requirements with >95% similarity'},
            {'Section': '', 'Metric': 'Duplicate Groups', 'Value': df[df['Duplicate_Group'] != 'UNIQUE']['Duplicate_Group'].nunique() if dup_count > 0 else 0, 'Details': ''},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # Methodology
        summary_data.extend([
            {'Section': 'SCORING METHODOLOGY', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Quality Score Calculation', 'Value': '', 'Details': 'Average of 5 dimensions (Clarity, Completeness, Verifiability, Atomicity, Consistency)'},
            {'Section': '', 'Metric': 'Issue Severity Weights', 'Value': '', 'Details': 'Critical: -10, High: -5, Medium: -2, Low: -1 points'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # Quality Dimensions
        summary_data.extend([
            {'Section': 'QUALITY DIMENSIONS', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Clarity', 'Value': '', 'Details': 'Uses unambiguous language, active voice, and concise structure so the intent is immediately understood without interpretation.'},
            {'Section': '', 'Metric': 'Completeness', 'Value': '', 'Details': 'Explicitly states all necessary information, including actors, actions, objects, conditions, and performance criteria so the requirement is fully defined.'},
            {'Section': '', 'Metric': 'Verifiability', 'Value': '', 'Details': 'States measurable acceptance criteria such that the requirements satisfaction can be proven by test, analysis, inspection, or demonstration.'},
            {'Section': '', 'Metric': 'Atomicity', 'Value': '', 'Details': 'Specifies only a single behavior or constraint so it does not combine multiple requirements into one statement.'},
            {'Section': '', 'Metric': 'Consistency', 'Value': '', 'Details': 'Maintains uniform terminology and correct use of modal verbs so it does not conflict with other requirements in the specification.'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # INCOSE Patterns
        summary_data.extend([
            {'Section': 'INCOSE PATTERNS', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Functional/Performance', 'Value': '', 'Details': 'The {AGENT} shall {FUNCTION} in accordance with {INTERFACE_OUTPUT} with {PERFORMANCE} [and {TIMING} upon {EVENT_TRIGGER} in accordance with {INTERFACE_INPUT}] while in {CONDITION}'},
            {'Section': '', 'Metric': 'Suitability', 'Value': '', 'Details': 'The {AGENT} shall exhibit {CHARACTERISTIC} with {PERFORMANCE} while {CONDITION} [for {CONDITION_DURATION}]'},
            {'Section': '', 'Metric': 'Environmental', 'Value': '', 'Details': 'The {AGENT} shall exhibit {CHARACTERISTIC} during/after exposure to {ENVIRONMENT} [for {EXPOSURE_DURATION}]'},
            {'Section': '', 'Metric': 'Design Constraint', 'Value': '', 'Details': 'The {AGENT} shall exhibit {DESIGN_CONSTRAINTS} [in accordance with {PERFORMANCE} while in {CONDITION}]'},
            {'Section': '', 'Metric': '', 'Value': '', 'Details': ''},
        ])
        
        # Semantic Checks
        summary_data.extend([
            {'Section': 'SEMANTIC ANALYSIS', 'Metric': '', 'Value': '', 'Details': ''},
            {'Section': '', 'Metric': 'Ambiguous Terms', 'Value': '', 'Details': 'Words like: appropriate, various, several, many, some'},
            {'Section': '', 'Metric': 'Entity Extraction', 'Value': '', 'Details': 'Identifies actors (who), actions (what), objects (on what)'},
            {'Section': '', 'Metric': 'Implementation Details', 'Value': '', 'Details': 'Detects "how" vs "what": using, via, through + technology'},
            {'Section': '', 'Metric': 'Similarity Threshold', 'Value': '95%', 'Details': 'For duplicate detection'},
        ])
        
        return pd.DataFrame(summary_data)
    
    def _get_grade_description(self, grade: str) -> str:
        """Get description for grade."""
        descriptions = {
            'EXCELLENT': 'Professional quality, ready for implementation',
            'GOOD': 'Minor issues only, acceptable for most uses',
            'FAIR': 'Several issues, needs improvement',
            'POOR': 'Significant issues, requires major revision',
            'CRITICAL': 'Severe issues, needs complete rewrite'
        }
        return descriptions.get(grade, '')
    
    def _print_summary(self, df: pd.DataFrame):
        """Print analysis summary."""
        print("\n" + "="*70)
        print("REQUIREMENTS QUALITY ANALYSIS SUMMARY")
        print("="*70)
        print(f"\n📊 Overall Statistics:")
        print(f"  Total Requirements: {len(df)}")
        print(f"  Average Quality Score: {df['Quality_Score'].mean():.1f}/100")
        print(f"  Average INCOSE Compliance: {df['INCOSE_Compliance_Score'].mean():.1f}%")
        print(f"  Average Semantic Score: {df['Semantic_Quality_Score'].mean():.1f}/100")
        
        print(f"\n🎯 Grade Distribution:")
        grades = df['Quality_Grade'].value_counts()
        for grade in ['EXCELLENT', 'GOOD', 'FAIR', 'POOR', 'CRITICAL']:
            count = grades.get(grade, 0)
            pct = count / len(df) * 100 if len(df) > 0 else 0
            print(f"  {grade}: {count} ({pct:.1f}%)")
        
        print(f"\n📋 Key Findings:")
        print(f"  Requirements with issues: {len(df[df['Total_Issues'] > 0])}")
        print(f"  Missing actors: {len(df[df['Actors_Found'] == 'None'])}")
        print(f"  Missing modal verbs: {len(df[df['Has_Agent'] == 'No'])}")
        print(f"  Potential duplicates: {len(df[df['Duplicate_Group'] != 'UNIQUE'])}")
        
        print("="*70)


def main():
    """Main function."""
    parser = argparse.ArgumentParser(
        description="Enhanced Requirements Quality Analyzer with 4-Tab Excel Output",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python reqGrading.py                    # Analyze requirements.csv
  python reqGrading.py -i my_reqs.csv    # Analyze custom file
  python reqGrading.py -c "Requirement"   # Use different column name

Output:
  • CSV file with all analysis results
  • Excel workbook with 4 tabs:
    - Summary: Overview and methodology
    - Quality Grading: Traditional quality metrics
    - INCOSE Analysis: Pattern compliance
    - Semantic Analysis: Entities and similarity
        """
    )
    
    parser.add_argument("-i", "--input", dest="input_file", default="requirements.csv",
                        help="Input CSV file with requirements")
    parser.add_argument("-o", "--output", dest="output_file", default=None,
                        help="Output filename (without extension)")
    parser.add_argument("-c", "--column", dest="requirement_column", default="Requirement Text",
                        help="Column name containing requirement text")
    parser.add_argument("-m", "--model", dest="spacy_model", default="en_core_web_lg",
                        help="spaCy model to use")
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="Enable verbose logging")
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    print("="*70)
    print("🚀 ENHANCED REQUIREMENTS QUALITY ANALYZER v2.0")
    print("="*70)
    
    # Create analyzer
    analyzer = EnhancedRequirementAnalyzer(spacy_model=args.spacy_model)
    
    # Run analysis
    try:
        results_df = analyzer.analyze_file(
            input_file=args.input_file,
            output_file=args.output_file,
            requirement_column=args.requirement_column,
            excel_report=True
        )
        
        print(f"\n✅ Analysis complete!")
        output_name = args.output_file or Path(args.input_file).stem + "_quality_analysis"
        print(f"📁 CSV results: {output_name}.csv")
        print(f"📊 Excel report: {output_name}.xlsx")
        
    except Exception as e:
        logger.error(f"Analysis failed: {e}")
        raise


if __name__ == "__main__":
    main()