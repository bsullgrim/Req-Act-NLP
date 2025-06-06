"""
Matching Workbook Generator - Create comprehensive Excel workbook for engineering review
Organizes matches by confidence level with discovery analysis and action items
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import logging

logger = logging.getLogger(__name__)

class MatchingWorkbookGenerator:
    """Generate comprehensive Excel workbook for engineering teams to review matches."""
    
    def __init__(self, repo_manager=None):
        self.workbook_path = None
        if repo_manager is None:
         raise ValueError("Repository manager is required")
        self.repo_manager = repo_manager
        
    def create_workbook(self, 
                    enhanced_df: pd.DataFrame,
                    evaluation_results: Optional[Dict] = None,
                    output_path: Optional[str] = None,
                    repo_manager=None) -> str:
        
        # Setup repository manager
        if repo_manager is None:
            from src.utils.repository_setup import RepositoryStructureManager
            repo_manager = RepositoryStructureManager("outputs")
            repo_manager.setup_repository_structure()
        
        # Use proper output path
        if output_path is None:
            output_path = repo_manager.structure['engineering_review'] / "matching_workbook.xlsx"
        else:
            output_path = Path(output_path)
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # Tab 1: Executive Summary
            exec_summary = self._create_executive_summary(enhanced_df, evaluation_results)
            exec_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
            
            # Tab 2-4: Confidence-based tabs
            self._create_confidence_tabs(enhanced_df, writer)
            
            # Tab 5: Discovery Analysis (if available)
            if evaluation_results and 'discovery_analysis' in evaluation_results:
                self._create_discovery_tab(evaluation_results['discovery_analysis'], writer)
            
            # Tab 6: Action Items
            action_items = self._create_action_items(enhanced_df)
            action_items.to_excel(writer, sheet_name='Action Items', index=False)
            
            # Format all worksheets
            self._format_workbook(writer)
        
        logger.info(f"✅ Matching workbook created: {output_path}")
        self.workbook_path = str(output_path)
        return str(output_path)
    
    def _create_executive_summary(self, enhanced_df: pd.DataFrame, 
                                evaluation_results: Optional[Dict]) -> pd.DataFrame:
        """Create executive summary with key metrics."""
        
        total_matches = len(enhanced_df)
        unique_reqs = enhanced_df['ID'].nunique()
        
        # Confidence distribution
        high_conf = len(enhanced_df[enhanced_df['Combined Score'] >= 0.8])
        med_conf = len(enhanced_df[(enhanced_df['Combined Score'] >= 0.5) & 
                                 (enhanced_df['Combined Score'] < 0.8)])
        low_conf = len(enhanced_df[enhanced_df['Combined Score'] < 0.5])
        
        summary_data = [
            # Overview metrics
            {"Category": "Overview", "Metric": "Total Matches", "Value": total_matches, 
             "Percentage": "100.0%", "Notes": "All algorithm suggestions"},
            {"Category": "Overview", "Metric": "Requirements Covered", "Value": unique_reqs,
             "Percentage": f"{unique_reqs/total_matches*100:.1f}%", "Notes": "Unique requirements with matches"},
            
            # Confidence distribution
            {"Category": "Confidence", "Metric": "High Confidence (≥0.8)", "Value": high_conf,
             "Percentage": f"{high_conf/total_matches*100:.1f}%", "Notes": "Ready for approval"},
            {"Category": "Confidence", "Metric": "Medium Confidence (0.5-0.8)", "Value": med_conf,
             "Percentage": f"{med_conf/total_matches*100:.1f}%", "Notes": "Needs review"},
            {"Category": "Confidence", "Metric": "Low Confidence (<0.5)", "Value": low_conf,
             "Percentage": f"{low_conf/total_matches*100:.1f}%", "Notes": "Manual investigation"},
            
            # Quality distribution (if available)
        ]
        
        if 'Quality_Grade' in enhanced_df.columns:
            quality_counts = enhanced_df['Quality_Grade'].value_counts()
            for grade, count in quality_counts.items():
                summary_data.append({
                    "Category": "Quality", "Metric": f"Requirements: {grade}", 
                    "Value": count, "Percentage": f"{count/total_matches*100:.1f}%",
                    "Notes": "Quality of underlying requirements"
                })
        
        # Evaluation metrics (if available)
        if evaluation_results and 'aggregate_metrics' in evaluation_results:
            agg_metrics = evaluation_results['aggregate_metrics']
            if 'f1_at_5' in agg_metrics:
                f1_score = agg_metrics['f1_at_5']['mean']
                summary_data.append({
                    "Category": "Performance", "Metric": "F1@5 Score", 
                    "Value": f"{f1_score:.3f}", "Percentage": f"{f1_score*100:.1f}%",
                    "Notes": "Algorithm validation against manual traces"
                })
        
        # Discovery metrics (if available)
        if evaluation_results and 'discovery_analysis' in evaluation_results:
            discovery = evaluation_results['discovery_analysis']['summary']
            summary_data.append({
                "Category": "Discovery", "Metric": "Novel Connections Found",
                "Value": discovery.get('total_high_scoring_misses', 0),
                "Percentage": f"{discovery.get('discovery_rate', 0)*100:.1f}%",
                "Notes": "High-scoring matches not in manual traces"
            })
        
        return pd.DataFrame(summary_data)
    
    def _create_confidence_tabs(self, enhanced_df: pd.DataFrame, writer):
        """Create separate tabs for each confidence level."""
        
        confidence_levels = [
            ("High Confidence (≥0.8)", enhanced_df[enhanced_df['Combined Score'] >= 0.8]),
            ("Medium Confidence (0.5-0.8)", enhanced_df[(enhanced_df['Combined Score'] >= 0.5) & 
                                                       (enhanced_df['Combined Score'] < 0.8)]),
            ("Low Confidence (<0.5)", enhanced_df[enhanced_df['Combined Score'] < 0.5])
        ]
        
        for tab_name, df_subset in confidence_levels:
            if len(df_subset) > 0:
                # Select and order columns for engineering review
                review_columns = self._get_review_columns(df_subset)
                review_df = df_subset[review_columns].copy()
                
                # Sort by score (highest first)
                review_df = review_df.sort_values('Combined Score', ascending=False)
                
                # Add review columns
                review_df['Review Status'] = 'PENDING'
                review_df['Assigned Engineer'] = ''
                review_df['Review Notes'] = ''
                review_df['Approval Decision'] = ''
                
                # Write to Excel
                review_df.to_excel(writer, sheet_name=tab_name, index=False)
                
                logger.info(f"   ✓ Created {tab_name} tab: {len(review_df)} matches")
    
    def _get_review_columns(self, df: pd.DataFrame) -> List[str]:
        """Get appropriate columns for engineering review."""
        
        base_columns = [
            'ID', 'Requirement Name', 'Requirement Text', 'Activity Name', 
            'Combined Score'
        ]
        
        # Add score components if available
        score_columns = ['Dense Semantic', 'BM25 Score', 'Syntactic Score', 
                        'Domain Weighted', 'Query Expansion']
        available_score_cols = [col for col in score_columns if col in df.columns]
        base_columns.extend(available_score_cols)
        
        # Add quality information if available
        quality_columns = ['Quality_Grade', 'Quality_Score']
        available_quality_cols = [col for col in quality_columns if col in df.columns]
        base_columns.extend(available_quality_cols)
        
        # Only return columns that exist in the DataFrame
        return [col for col in base_columns if col in df.columns]
    
    def _create_discovery_tab(self, discovery_analysis: Dict, writer):
        """Create discovery analysis tab if evaluation results available."""
        
        high_scoring_misses = discovery_analysis.get('high_scoring_misses', [])
        
        if high_scoring_misses:
            discovery_df = pd.DataFrame(high_scoring_misses)
            
            # Add investigation columns
            discovery_df['Investigation Status'] = 'PENDING'
            discovery_df['Assigned Investigator'] = ''
            discovery_df['Investigation Notes'] = ''
            discovery_df['Outcome'] = ''
            
            # Sort by score
            discovery_df = discovery_df.sort_values('score', ascending=False)
            
            discovery_df.to_excel(writer, sheet_name='Discovery Analysis', index=False)
            logger.info(f"   ✓ Created Discovery Analysis tab: {len(discovery_df)} discoveries")
    
    def _create_action_items(self, enhanced_df: pd.DataFrame) -> pd.DataFrame:
        """Create prioritized action items for project management."""
        
        action_items = []
        
        # Group by confidence level for different action types
        high_conf = enhanced_df[enhanced_df['Combined Score'] >= 0.8]
        med_conf = enhanced_df[(enhanced_df['Combined Score'] >= 0.5) & 
                              (enhanced_df['Combined Score'] < 0.8)]
        low_conf = enhanced_df[enhanced_df['Combined Score'] < 0.5]
        
        # High confidence - approval workflow
        for _, row in high_conf.head(20).iterrows():  # Top 20 for manageable list
            action_items.append({
                'Priority': 'HIGH',
                'Action Type': 'APPROVE',
                'Requirement ID': row['ID'],
                'Requirement Name': row.get('Requirement Name', 'N/A'),
                'Activity Name': row['Activity Name'],
                'Score': row['Combined Score'],
                'Assigned To': '',
                'Due Date': '',
                'Status': 'PENDING',
                'Notes': f'High confidence match ({row["Combined Score"]:.3f}) - fast-track approval'
            })
        
        # Medium confidence - detailed review
        for _, row in med_conf.head(15).iterrows():  # Top 15
            action_items.append({
                'Priority': 'MEDIUM',
                'Action Type': 'REVIEW',
                'Requirement ID': row['ID'],
                'Requirement Name': row.get('Requirement Name', 'N/A'),
                'Activity Name': row['Activity Name'],
                'Score': row['Combined Score'],
                'Assigned To': '',
                'Due Date': '',
                'Status': 'PENDING',
                'Notes': f'Medium confidence ({row["Combined Score"]:.3f}) - needs detailed review'
            })
        
        # Low confidence - investigation
        for _, row in low_conf.head(10).iterrows():  # Top 10
            action_items.append({
                'Priority': 'LOW',
                'Action Type': 'INVESTIGATE',
                'Requirement ID': row['ID'],
                'Requirement Name': row.get('Requirement Name', 'N/A'),
                'Activity Name': row['Activity Name'],
                'Score': row['Combined Score'],
                'Assigned To': '',
                'Due Date': '',
                'Status': 'PENDING',
                'Notes': f'Low confidence ({row["Combined Score"]:.3f}) - manual investigation'
            })
        
        return pd.DataFrame(action_items)
    
    def _format_workbook(self, writer):
        """Apply formatting to the workbook."""
        
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Get the workbook and worksheets
        workbook = writer.book
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        high_conf_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        med_conf_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        low_conf_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Format headers
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply confidence-based row coloring for confidence tabs
            if "Confidence" in sheet_name and "Combined Score" in [cell.value for cell in worksheet[1]]:
                score_col = None
                for idx, cell in enumerate(worksheet[1], 1):
                    if cell.value == "Combined Score":
                        score_col = idx
                        break
                
                if score_col:
                    for row in worksheet.iter_rows(min_row=2):
                        score_value = row[score_col-1].value
                        if isinstance(score_value, (int, float)):
                            if score_value >= 0.8:
                                fill = high_conf_fill
                            elif score_value >= 0.5:
                                fill = med_conf_fill
                            else:
                                fill = low_conf_fill
                            
                            for cell in row:
                                cell.fill = fill

def create_matching_workbook(enhanced_df: pd.DataFrame, 
                           evaluation_results: Optional[Dict] = None,
                           output_path: str = "outputs/engineering_review/matching_workbook.xlsx") -> str:
    """
    Convenience function to create matching workbook.
    
    Args:
        enhanced_df: Enhanced predictions DataFrame
        evaluation_results: Optional evaluation results with discovery analysis
        output_path: Where to save the workbook
        
    Returns:
        Path to created workbook
    """
    generator = MatchingWorkbookGenerator()
    return generator.create_workbook(enhanced_df, evaluation_results, output_path)